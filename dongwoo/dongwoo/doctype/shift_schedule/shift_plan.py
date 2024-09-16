import frappe
import frappe.utils
from frappe.utils.csvutils import read_csv_content
from frappe.utils import get_first_day, get_last_day, format_datetime, get_url_to_form
from frappe.utils import cint
from frappe.utils.data import date_diff, now_datetime, nowdate, today, add_days
import datetime
from io import BytesIO
import openpyxl
from frappe import _
from frappe import _
from frappe.model.document import Document
from frappe.utils import getdate, nowdate
from frappe import throw, msgprint
import frappe
from frappe.utils import flt, fmt_money
from datetime import timedelta
from datetime import date
from frappe import throw, _
from frappe.utils import getdate, today
today = date.today()
from frappe.model.document import Document
import datetime 
import frappe,erpnext
from frappe.utils import cint
import json
from frappe.utils import date_diff, add_months,today,add_days,add_years,nowdate,flt
from frappe.model.mapper import get_mapped_doc
from frappe.utils.file_manager import get_file
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
import datetime
from datetime import date,datetime,timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from frappe.utils import formatdate
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from io import BytesIO
from datetime import datetime, timedelta
from frappe.utils import now
from datetime import date
from six import BytesIO, string_types
from frappe.utils import time_diff

@frappe.whitelist()
def shift_plan_excel():
    filename = "Shift" + today()
    xlsx_file = build_xlsx_response(filename)
    send_mail_with_attachment(filename, xlsx_file.getvalue())


def send_mail_with_attachment(filename, file_content):
    subject = ("Shift-%s"%(nowdate()) )
    message = "Dear Sir/Madam,<br> Please find attached Report.<br>Thanks & Regards,<br>TEAM ERP<br>This email has been automatically generated. Please do not reply"
    attachments = [{"fname": filename + '.xlsx', "fcontent": file_content}]
    frappe.sendmail(
		recipients= ['pavithra.s@groupteampro.com',"venkatrajr@dwsi.co.in",'vishal@dwsi.co.in',"vinothkumar@dwsi.co.in","vishnu@dwsi.co.in","security@dwsi.co.in"],
        sender=None,  
        subject=subject,
        message=message,
        attachments=attachments,
    )


def build_xlsx_response(filename):
    xlsx_file = make_xlsx(filename)
    return xlsx_file


def make_xlsx(filename, sheet_name=None, wb=None, column_widths=None):
    column_widths = column_widths or []
    if wb is None:
        wb = Workbook()
    ws = wb.create_sheet(sheet_name or "Shift Assignments", 0)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)    
    end_date = datetime.now().date()
    start_date = end_date - timedelta(days=7)  
    start_date_str = start_date.strftime('%Y-%m-%d')
    end_date_str = (end_date - timedelta(days=1)).strftime('%Y-%m-%d')  
    header_title = f"Weekly Attendance Report ({start_date_str} - {end_date_str})"
    merged_cell = ws.cell(row=1, column=1, value=header_title)
    merged_cell.font = Font(bold=True)
    merged_cell.alignment = Alignment(horizontal="center", vertical="center")
    headers = ["SI NO", "Department", "Schedule", "Actual", "Shortage", "Contractor Scope OT to Manage shortage", "Extra Manpower (OT) support to Manage D.Trainee / Operator shortage Company OT"]
    ws.append(headers)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = header
        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
        cell.border = Border(left=Side(border_style='thin', color='000000'),right=Side(border_style='thin', color='000000'),top=Side(border_style='thin', color='000000'),bottom=Side(border_style='thin', color='000000'))
    default_column_widths = [10, 25, 25, 25, 25, 25, 25]
    column_widths = column_widths or default_column_widths    
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width  
    dept = ['Production', 'Maintenance','QC', 'PMT','PE']
    si_no = 0
    data_rows = []
    for i in dept:
        si_no += 1
        if i=='PMT':
            shift_count = frappe.db.count("Shift Assignment", {"department": ["in", ['D.PMT','H.PMT']],"start_date": ["between", [start_date_str, end_date_str]],"docstatus": ("!=", 2)})
            att_count = frappe.db.count("Attendance", {"department": ["in", ['D.PMT','H.PMT']],"attendance_date": ["between", [start_date_str, end_date_str]],"status": "Present","docstatus": ("!=", 2)})
            ot_count = frappe.db.count("Attendance", {"department": ["in", ['D.PMT','H.PMT']],"attendance_date": ["between", [start_date_str, end_date_str]],"status": "Present","overtime_hours": [">=", '8'], "docstatus": ("!=", 2)})
        else:
            shift_count = frappe.db.count("Shift Assignment", {"department":  i,"start_date": ["between", [start_date_str, end_date_str]],"docstatus": ("!=", 2)})
            att_count = frappe.db.count("Attendance", {"department": i,"attendance_date": ["between", [start_date_str, end_date_str]],"status": "Present","docstatus": ("!=", 2)})
            ot_count = frappe.db.count("Attendance", {"department":  i,"attendance_date": ["between", [start_date_str, end_date_str]],"status": "Present","overtime_hours": [">=", '8'], "docstatus": ("!=", 2)})
        
        shortage = shift_count-att_count
        ot_shortage = shortage-ot_count
        data_rows.append([si_no, i, shift_count, att_count, shortage, ot_count, ot_shortage])
    for row in data_rows:
        ws.append(row)
    totals_row = ["Total", "", 0, 0, 0, 0, 0]
    for col_idx in range(3, len(headers) + 1):  
        totals_row[col_idx - 1] = sum(row[col_idx - 1] for row in data_rows)
    ws.append(totals_row)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
            cell.border = Border(left=Side(border_style='thin', color='000000'),right=Side(border_style='thin', color='000000'),top=Side(border_style='thin', color='000000'),bottom=Side(border_style='thin', color='000000'))
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=ws.max_row, column=col_idx)
        cell.fill = PatternFill(start_color="45b509", end_color="45b509", fill_type="solid")
        cell.font = cell.font.copy(bold=True) 
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center") 
        cell.border = Border(left=Side(border_style='thin', color='000000'),right=Side(border_style='thin', color='000000'),top=Side(border_style='thin', color='000000'),bottom=Side(border_style='thin', color='000000'))    
    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    xlsx_file.seek(0) 
    return xlsx_file

def shift_plan_count():
    job = frappe.db.exists('Scheduled Job Type', 'shift_plan_excel')
    if not job:
        var = frappe.new_doc("Scheduled Job Type")
        var.update({
            "method": 'dongwoo.dongwoo.doctype.shift_schedule.shift_plan.shift_plan_excel',
            "frequency": 'Cron',
            "cron_format": '30 11 * * 2'
        })
        var.save(ignore_permissions=True) 