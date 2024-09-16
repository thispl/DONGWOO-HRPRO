import frappe
from frappe.utils.background_jobs import enqueue
from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from six import BytesIO
from datetime import datetime, timedelta
import calendar
from frappe.utils import getdate

@frappe.whitelist()
def download():
    filename = 'Monthly Salary Report'
    test = build_xlsx_response(filename)

   
    

def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
    args = frappe.local.form_dict
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()

    att_date = getdate(args.start_date)
    att_month = att_date.strftime("%B")
    att_year = att_date.year
    # Execute the SQL query to fetch employee details along with corresponding salary slip names
    employee_details = frappe.db.sql("""
    SELECT e.bank_ac_no AS 'AC No', e.name, e.first_name AS 'Employee Name', s.employee_name AS 'Salary Slip Ename', e.mop, e.employee_type AS 'Employee Type',
           e.designation AS 'Designation', e.employee_type AS 'Etype', e.department AS 'Department', e.uan_number AS "UAN NO", e.esi_number AS "Esi", e.date_of_joining AS 'Date of Joining',
           s.payment_days AS 'Payment Days', s.leave_without_pay AS 'LOP', s.total_working_days AS 'Days in Month',
           e.basic AS 'Basic Pay', e.house_rent_allowance AS 'HRA', e.medical_allowance AS 'Medical Allowance', e.conveyance_allowance AS 'Conveyance',
           e.education_allowance AS 'Education Allowance', e.leave_and_travel_allowance AS 'LTA', e.dress_allowance AS 'Dress Allowance', e.gross_pay AS 'Fixed Gross', 
           s.name AS 'Name', s.total_deduction AS 'Total Deduction'
    FROM `tabEmployee` e
    INNER JOIN `tabSalary Slip` s ON e.name = s.employee
    INNER JOIN `tabDepartment` d ON e.department = d.name
    WHERE e.name = s.employee AND s.start_date <= %s AND s.end_date >= %s
    ORDER BY d.order_value, e.employee_type, e.date_of_joining
""", (args.start_date, args.end_date), as_dict=True)# frappe.errprint(args.start_date)
    # frappe.errprint(args.end_date)
    start_date = datetime.strptime(args.start_date, '%Y-%m-%d')
    current_year = start_date.year
    current_month = start_date.month
    if current_month == 1:
        prev_month = 12
        prev_year = current_year - 1
    else:
        prev_month = current_month - 1
        prev_year = current_year
    if prev_month == 1:
        prev_prev_month = 12
        prev_prev_year = prev_year - 1
    else:
        prev_prev_month = prev_month - 1
        prev_prev_year = prev_year
    prev_month_name = calendar.month_name[prev_month]
    next_month_name = calendar.month_name[prev_prev_month]
    pmonth= 'OT'+'('+prev_month_name+')'
    nmonth= 'OT'+'('+next_month_name+')'
    if start_date.month == 1:
        prev_month = 12
        prev_year = start_date.year - 1
    else:
        prev_month = start_date.month - 1
        prev_year = start_date.year
    
    prev_month_start = datetime(prev_year, prev_month, 1)
    prev_month_end = (prev_month_start + timedelta(days=31)).replace(day=1) - timedelta(days=1)    
    if prev_month == 1:
        prev_prev_month = 12
        prev_prev_year = prev_year - 1
    else:
        prev_prev_month = prev_month - 1
        prev_prev_year = prev_year
    prev_prev_month_start = datetime(prev_prev_year, prev_prev_month, 1)
    prev_prev_month_end = (prev_prev_month_start + timedelta(days=31)).replace(day=1) - timedelta(days=1)
    # prev_month_start, prev_month_end, prev_prev_month_start, prev_prev_month_end
    ws = wb.create_sheet(sheet_name, 0)
    ws.append(['DWSI SALARY STATEMENT FOR THE MONTH OF' ''  +str(att_month)+ '-' +str(att_year)])
    
    ws.append(['S.NO','DEPT','EMP NO.','NAMES','A/c No','MOP','DESIGNATION','PF UAN NO','ESI NO','DATE OF JOINING','FIXED GROSS','','','','','','','','DAYS','','','SALARY CALCULATION','','','','','','','','','','','','','','','','','','','','DEDUCTION','','','','','','','','','','','TAKE HOME'])

    ws.append(['S.NO','DEPT','EMP NO.','NAMES','A/c No','MOP','DESIGNATION','PF UAN NO','ESI NO','DATE OF JOINING','Basic Pay','House rent Allowance','Medical Allowance','Convey.Allow','Edu.Allow','Leave Travel Allow','Dress Allow','Fixed Gross','LOP Days','No.of.Days Paid','Days in a Month','Basic Pay','House rent Allowance','Medical Allowance','Convey.Allow','Edu.Allow','Leave Travel Allow','Dress Allow','Gross','OT',pmonth,nmonth,'Bus Fare','Fest.Allow','Arrears','Attendance Bonus','Supervisor Allow','Performance Allow','Shift Allow','Special Duty Allow','Total','PF(12%)','VPF','TOTAL PF DED.','LWF','IT','Prof.Tax','Advance','Miscellenous Ded','ESI','Total Deduction','LOP'])
    
    
    serial_number = 1
    previous_department = None
    department_total = 0
    total=0
    sub=0
    count=0
    count_1=0
    total1=0
    total_pf_dec=0
    takehome=0
    
    grand_total_basic=0
    grand_total_hra=0
    grand_total_med=0
    grand_total_convey=0
    grand_total_edu=0
    grand_total_lta=0
    grand_total_dress=0
    grand_total_gross=0
    grand_total_lop=0
    grand_total_days=0
    grand_total_month=0
    grand_total_ebasic=0
    grand_total_ehra=0
    grand_total_emed=0
    grand_total_elwf=0
    grand_total_econvey=0
    grand_total_eedu=0
    grand_total_elta=0
    grand_total_edress=0
    grand_total_ot=0
    grand_total_egross=0
    grand_total_bus=0
    grand_total_fes=0
    grand_total_arrear=0
    grand_total_att=0
    grand_total_supervisor=0
    grand_total_shift=0
    grand_total_spec=0
    grand_total_etotal=0
    grand_total_pf=0
    grand_total_vpf=0
    grand_total_totalpf=0
    grand_total_it=0
    grand_total_mis=0
    grand_total_pt=0
    grand_total_pot=0
    grand_total_ppot=0
    grand_total_advance=0
    grand_total_esi=0
    grand_total_totaldec=0
    grand_total_dlop=0
    grand_total_takehome=0
    grand_total_desig=0
    wgrand_total_basic=0
    wgrand_total_hra=0
    wgrand_total_med=0
    wgrand_total_convey=0
    wgrand_total_edu=0
    wgrand_total_lta=0
    wgrand_total_dress=0
    wgrand_total_gross=0
    wgrand_total_lop=0
    wgrand_total_days=0
    wgrand_total_month=0
    wgrand_total_ebasic=0
    wgrand_total_elwf=0
    wgrand_total_ehra=0
    wgrand_total_emed=0
    wgrand_total_econvey=0
    wgrand_total_eedu=0
    wgrand_total_elta=0
    wgrand_total_edress=0
    wgrand_total_ot=0
    wgrand_total_egross=0
    wgrand_total_bus=0
    wgrand_total_fes=0
    wgrand_total_arrear=0
    wgrand_total_att=0
    wgrand_total_supervisor=0
    wgrand_total_shift=0
    wgrand_total_spec=0
    wgrand_total_etotal=0
    wgrand_total_pf=0
    wgrand_total_vpf=0
    wgrand_total_totalpf= 0
    wgrand_total_it=0
    wgrand_total_mis=0
    wgrand_total_pt=0
    wgrand_total_pot=0
    wgrand_total_ppot=0
    wgrand_total_advance=0
    wgrand_total_esi=0
    wgrand_total_totaldec=0
    wgrand_total_dlop=0
    wgrand_total_takehome=0
    wgrand_total_desig=0
    grand_total_desigw=0
    # Append data to worksheet
    for row in employee_details:
        frappe.errprint(row['name'])
        prevslip = frappe.db.get_value("Salary Slip", {"employee": row['name'] ,"start_date":prev_month_start,"end_date":prev_month_end},["name"]) 
        if prevslip:
            prevot = frappe.db.get_value("Salary Detail", {"parent": prevslip , "salary_component": "Overtime"},"amount") or 0.0
        else:
            prevot = 0.0
        pprevslip = frappe.db.get_value("Salary Slip", {"employee": row['name'] ,"start_date":prev_prev_month_start,"end_date":prev_prev_month_end},["name"]) 
        if pprevslip:
            pprevot = frappe.db.get_value("Salary Detail", {"parent": pprevslip , "salary_component": "Overtime"},"amount") or 0.0
        else:
            pprevot = 0.0
        ebasic = frappe.db.get_value("Salary Detail", {"parent": row['Name'] , "salary_component": "Basic"},"amount") or 0.0
        ehra = frappe.db.get_value("Salary Detail", {"parent": row['Name'] , "salary_component": "House Rent Allowance"},"amount") or 0.0
        emedical= frappe.db.get_value("Salary Detail", {"parent": row['Name'] , "salary_component": "Medical Allowance"},"amount") or 0.0
        econveyance = frappe.db.get_value("Salary Detail", {"parent": row['Name'] , "salary_component": "Conveyance Allowance"},"amount")or 0.0
        eeducation = frappe.db.get_value("Salary Detail", {"parent": row['Name'] , "salary_component": "Education Allowance"},"amount") or 0.0
        elta = frappe.db.get_value("Salary Detail", {"parent": row['Name'] , "salary_component": "Leave and Travel Allowance"},"amount") or 0.0
        edress = frappe.db.get_value("Salary Detail", {"parent": row['Name'] , "salary_component": "Dress Allowance"},"amount") or 0.0   
        eot = frappe.db.get_value("Salary Detail", {"parent": row['Name'] , "salary_component": "Overtime"},"amount") or 0.0
        ebus = frappe.db.get_value("Salary Detail", {"parent": row['Name'] , "salary_component": "Bus Fare"},"amount") or 0.0
        efes = frappe.db.get_value("Salary Detail", {"parent": row['Name'] , "salary_component": "Festival Allowance"},"amount") or 0.0
        earrear = frappe.db.get_value("Salary Detail", {"parent": row['Name'] , "salary_component": "Arrear"},"amount") or 0.0
        eattendance = frappe.db.get_value("Salary Detail", {"parent": row['Name'] , "salary_component": "Attendance Bonus"},"amount") or 0.0
        esupervisor = frappe.db.get_value("Salary Detail", {"parent": row['Name'] , "salary_component": "Supervisor Allowance"},"amount") or 0.0
        eshift = frappe.db.get_value("Salary Detail", {"parent": row['Name'] , "salary_component": "Shift Allowance"},"amount") or 0.0
        esda = frappe.db.get_value("Salary Detail", {"parent": row['Name'] , "salary_component": "Special Duty Allowance"},"amount") or 0.0
        elwf = frappe.db.get_value("Salary Detail", {"parent": row['Name'] , "salary_component": "Labour Welfare  Fund"},"amount") or 0.0
        dprovident = frappe.db.get_value("Salary Detail", {"parent": row['Name'] , "salary_component": "Provident Fund"},"amount") or 0.0
        dvpf = frappe.db.get_value("Salary Detail", {"parent": row['Name'] , "salary_component": "Voluntary Provident Fund"},"amount") or 0.0
        dit = frappe.db.get_value("Salary Detail", {"parent": row['Name'] , "salary_component": "Income Tax"},"amount") or 0.0
        dpt = frappe.db.get_value("Salary Detail", {"parent": row['Name'] , "salary_component": "Professional Tax"},"amount") or 0.0
        dadvance = frappe.db.get_value("Salary Detail", {"parent": row['Name'] , "salary_component": "Advance"},"amount") or 0.0
        desi = frappe.db.get_value("Salary Detail", {"parent": row['Name'] , "salary_component": "Employee State Insurance"},"amount") or 0.0		
        dlop = frappe.db.get_value("Salary Detail", {"parent": row['Name'] , "salary_component": "Loss Of Pay"},"amount") or 0.0
        mis = frappe.db.get_value("Salary Detail", {"parent": row['Name'] , "salary_component": "Miscellenous"},"amount") or 0.0
        egross = ebasic+ehra+emedical+econveyance+eeducation+elta+edress
        total = egross +ebus+efes+earrear+eattendance+esupervisor+eshift+esda+eot
        takehome= total - row['Total Deduction']
        
        if row['Department'] == previous_department:
            # department_name = ''
            print(previous_department)
        else:
            if previous_department:
                count=0
                
                # Add subtotal row for the previous department
                ws.append(['','', 'Subtotal','', '', *sub, '', '','', '',sub_total_basic, sub_total_hra,sub_total_med, sub_total_convey,sub_total_edu, sub_total_lta, sub_total_dress, sub_total_gross, 
                sub_total_lop, sub_total_days, sub_total_month, sub_total_ebasic, sub_total_ehra,sub_total_emed, sub_total_econvey, sub_total_eedu, sub_total_elta, sub_total_edress,
                sub_total_egross,sub_total_ot,sub_total_pot,sub_total_ppot, sub_total_bus, sub_total_fes, sub_total_arrear, sub_total_att, sub_total_supervisor,0,sub_total_shift,sub_total_spec, sub_total_etotal, 
                sub_total_pf, sub_total_vpf, sub_total_totalpf,sub_total_elwf, sub_total_it, sub_total_pt, sub_total_advance,sub_total_mis,sub_total_esi,sub_total_totaldec, sub_total_dlop,sub_total_takehome])
            # total = egross+ebus+efes+earrear+eattendance+esupervisor+eshift+esda
            department_name = row['Department']
            previous_department = row['Department']
            sub_total_basic=0
            sub_total_hra = 0
            sub_total_med=0
            sub_total_convey=0
            sub_total_edu=0
            sub_total_lta=0
            sub_total_dress=0
            sub_total_gross=0
            sub_total_lop=0
            sub_total_elwf=0
            sub_total_days=0
            sub_total_month=0
            sub_total_ebasic=0
            sub_total_ehra=0
            sub_total_emed=0
            sub_total_econvey=0
            sub_total_eedu=0
            sub_total_elta=0
            sub_total_edress=0
            sub_total_ot=0
            sub_total_egross=0
            sub_total_bus=0
            sub_total_fes=0
            sub_total_arrear=0
            sub_total_att=0
            sub_total_supervisor=0
            sub_total_shift=0
            sub_total_spec=0
            sub_total_etotal=0
            sub_total_pf=0
            sub_total_vpf=0
            sub_total_totalpf=0
            sub_total_it=0
            sub_total_pot=0
            sub_total_ppot=0
            sub_total_mis=0
            sub_total_pt=0
            sub_total_advance=0
            sub_total_esi=0
            sub_total_totaldec=0
            sub_total_dlop=0
            sub_total_takehome=0

        total_pf_dec = dprovident+dvpf

        sub_total_basic+= row['Basic Pay']  
        sub_total_hra+=row['HRA']
        sub_total_med+=row['Medical Allowance']
        sub_total_convey+=row['Conveyance']
        sub_total_edu+=row['Education Allowance']
        sub_total_lta+=row['LTA']
        sub_total_dress+=row['Dress Allowance']
        sub_total_gross+=row['Fixed Gross']
        sub_total_lop+=row['LOP']
        sub_total_days+=row['Payment Days']
        sub_total_month+=row['Days in Month']
        sub_total_ebasic+=ebasic
        sub_total_ehra+=ehra
        sub_total_emed+=emedical
        sub_total_econvey+=econveyance
        sub_total_eedu+=eeducation
        sub_total_elta+=elta
        sub_total_edress+=edress
        sub_total_ot+=eot
        sub_total_elwf+=elwf
        sub_total_egross+=egross
        sub_total_bus+=ebus
        sub_total_fes+=efes
        sub_total_arrear+=earrear
        sub_total_att+=eattendance
        sub_total_supervisor+=esupervisor
        sub_total_shift+=eshift
        sub_total_spec+=esda
        sub_total_etotal+=total
        sub_total_pf+=dprovident
        sub_total_vpf+=dvpf
        sub_total_totalpf+= total_pf_dec
        sub_total_it+=dit
        sub_total_mis+=mis
        sub_total_pt+=dpt
        sub_total_pot+=prevot
        sub_total_ppot+=pprevot
        sub_total_advance+=dadvance
        sub_total_esi+=desi
        sub_total_totaldec+=row['Total Deduction']
        sub_total_lop+=dlop
        sub_total_takehome+=takehome
        
        grand_total_basic+=row['Basic Pay']
        grand_total_hra+=row['HRA']
        grand_total_med+=row['Medical Allowance']
        grand_total_convey+=row['Conveyance']
        grand_total_edu+=row['Education Allowance']
        grand_total_lta+=row['LTA']
        grand_total_dress+=row['Dress Allowance']
        grand_total_gross+=row['Fixed Gross']
        grand_total_lop+=row['LOP']
        grand_total_days+=row['Payment Days']
        grand_total_month+=row['Days in Month']
        grand_total_ebasic+=ebasic
        grand_total_ehra+=ehra
        grand_total_elwf+=elwf
        grand_total_emed+=emedical
        grand_total_econvey+=econveyance
        grand_total_eedu+=eeducation
        grand_total_elta+=elta
        grand_total_edress+=edress
        grand_total_ot+=eot
        grand_total_egross+=egross
        grand_total_bus+=ebus
        grand_total_fes+=efes
        grand_total_arrear+=earrear
        grand_total_att+=eattendance
        grand_total_supervisor+=esupervisor
        grand_total_shift+=eshift
        grand_total_spec+=esda
        grand_total_etotal+=total
        grand_total_pf+=dprovident
        grand_total_vpf+=dvpf
        grand_total_totalpf+= total_pf_dec
        grand_total_it+=dit
        grand_total_mis+=mis
        grand_total_pt+=dpt
        grand_total_pot+=prevot
        grand_total_ppot+=pprevot
        grand_total_advance+=dadvance
        grand_total_esi+=desi
        grand_total_totaldec+=row['Total Deduction']
        grand_total_dlop+=dlop
        grand_total_takehome+=takehome
        grand_total_desig=serial_number
        if row['Etype'] == 'Worker':
            grand_total_desigw+=1
            wgrand_total_basic+=row['Basic Pay']
            wgrand_total_hra+=row['HRA']
            wgrand_total_med+=row['Medical Allowance']
            wgrand_total_convey+=row['Conveyance']
            wgrand_total_edu+=row['Education Allowance']
            wgrand_total_lta+=row['LTA']
            wgrand_total_dress+=row['Dress Allowance']
            wgrand_total_gross+=row['Fixed Gross']
            wgrand_total_lop+=row['LOP']
            wgrand_total_days+=row['Payment Days']
            wgrand_total_month+=row['Days in Month']
            wgrand_total_ebasic+=ebasic
            wgrand_total_ehra+=ehra
            wgrand_total_emed+=emedical
            wgrand_total_econvey+=econveyance
            wgrand_total_eedu+=eeducation
            wgrand_total_elta+=elta
            wgrand_total_elwf+=elwf
            wgrand_total_edress+=edress
            wgrand_total_ot+=eot
            wgrand_total_egross+=egross
            wgrand_total_bus+=ebus
            wgrand_total_fes+=efes
            wgrand_total_arrear+=earrear
            wgrand_total_att+=eattendance
            wgrand_total_supervisor+=esupervisor
            wgrand_total_shift+=eshift
            wgrand_total_spec+=esda
            wgrand_total_etotal+=total
            wgrand_total_pf+=dprovident
            wgrand_total_vpf+=dvpf
            wgrand_total_totalpf+= total_pf_dec
            wgrand_total_it+=dit
            wgrand_total_mis+=mis
            wgrand_total_pt+=dpt
            wgrand_total_pot+=prevot
            wgrand_total_ppot+=pprevot
            wgrand_total_advance+=dadvance
            wgrand_total_esi+=desi
            wgrand_total_totaldec+=row['Total Deduction']
            wgrand_total_dlop+=dlop
            wgrand_total_takehome+=takehome
            wgrand_total_desig=serial_number
        
        if row['Etype'] == 'D . Trainee':
            ws.append([serial_number, department_name, row['name'], row['Employee Name'],row['AC No'],row['mop'], row['Designation'], row['UAN NO'],row['Esi'],row['Date of Joining'], row['Basic Pay'], row['HRA'], row['Medical Allowance'], row['Conveyance'], row['Education Allowance'], row['LTA'], row['Dress Allowance'], row['Fixed Gross'], row['LOP'], row['Payment Days'], row['Days in Month'], ebasic, ehra, emedical, econveyance, eeducation, elta, edress, egross,eot,prevot,pprevot, ebus, efes, earrear, eattendance, esupervisor,0, eshift, esda,total, dprovident, dvpf, total_pf_dec, elwf,dit, dpt, dadvance, mis,desi, row['Total Deduction'],dlop,takehome])
        else:
            ws.append([serial_number, department_name, row['name'], row['Employee Name'],row['AC No'],row['mop'], row['Designation'], row['UAN NO'],'-',row['Date of Joining'], row['Basic Pay'], row['HRA'], row['Medical Allowance'], row['Conveyance'], row['Education Allowance'], row['LTA'], row['Dress Allowance'], row['Fixed Gross'], row['LOP'], row['Payment Days'], row['Days in Month'], ebasic, ehra, emedical, econveyance, eeducation, elta, edress, egross,eot,prevot,pprevot,  ebus, efes, earrear, eattendance, esupervisor,0, eshift, esda,total, dprovident, dvpf, total_pf_dec, elwf,dit, dpt, dadvance, mis,desi,row['Total Deduction'], dlop,takehome])

        serial_number += 1
        count+=1
        sub = [str(count) + "MEMBERS"]
    
    
    mem= [str(grand_total_desig)+" MEMBERS"] 
    wmem= [str(grand_total_desigw)+" MEMBERS"] 
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=53)
    for i in range(1,11):
        ws.merge_cells(start_row=2,start_column=i,end_row=3,end_column=i)
        i=i+1
    ws.merge_cells(start_row=2,start_column=11,end_row=2,end_column=18)
    ws.merge_cells(start_row=2,start_column=19,end_row=2,end_column=21)
    ws.merge_cells(start_row=2,start_column=22,end_row=2,end_column=41)
    ws.merge_cells(start_row=2,start_column=42,end_row=2,end_column=52)
    ws.merge_cells(start_row=2,start_column=53,end_row=3,end_column=53)
    
    ws.append(['', 'Subtotal','Subtotal', '','', *sub, '', '', '','',sub_total_basic, sub_total_hra,sub_total_med, sub_total_convey,sub_total_edu, sub_total_lta, sub_total_dress, sub_total_gross, 
            sub_total_lop, sub_total_days, sub_total_month, sub_total_ebasic, sub_total_ehra,sub_total_emed, sub_total_econvey, sub_total_eedu, sub_total_elta, sub_total_edress,
                sub_total_egross, sub_total_ot,sub_total_pot,sub_total_ppot,sub_total_bus, sub_total_fes, sub_total_arrear, sub_total_att, sub_total_supervisor,0,sub_total_shift,sub_total_spec, sub_total_etotal, 
                sub_total_pf, sub_total_vpf, sub_total_totalpf,sub_total_elwf, sub_total_it, sub_total_pt, sub_total_advance,sub_total_mis, sub_total_esi, sub_total_totaldec, sub_total_dlop, sub_total_takehome])
    
    ws.append(['', 'Total for Workers','Total for Workers', '','',*wmem,'','','','',wgrand_total_basic, wgrand_total_hra, wgrand_total_med, wgrand_total_convey, wgrand_total_edu, wgrand_total_lta, wgrand_total_dress, wgrand_total_gross,
               wgrand_total_lop, wgrand_total_days, wgrand_total_month, wgrand_total_ebasic,wgrand_total_ehra,wgrand_total_emed, wgrand_total_econvey, wgrand_total_eedu, wgrand_total_elta, wgrand_total_edress,
                 wgrand_total_egross, wgrand_total_ot,wgrand_total_pot,wgrand_total_ppot,wgrand_total_bus, wgrand_total_fes, wgrand_total_arrear, wgrand_total_att, wgrand_total_supervisor,0,wgrand_total_shift,wgrand_total_spec, wgrand_total_etotal, 
                 wgrand_total_pf, wgrand_total_vpf, wgrand_total_totalpf, wgrand_total_elwf,wgrand_total_it, wgrand_total_pt, wgrand_total_advance,wgrand_total_mis, wgrand_total_esi, wgrand_total_totaldec, wgrand_total_dlop, wgrand_total_takehome])
    
    ws.append(['','Grand Total','Grand Total','','',*mem,'','','','',grand_total_basic, grand_total_hra, grand_total_med, grand_total_convey, grand_total_edu, grand_total_lta, grand_total_dress, grand_total_gross,
               grand_total_lop, grand_total_days, grand_total_month, grand_total_ebasic,grand_total_ehra,grand_total_emed, grand_total_econvey, grand_total_eedu, grand_total_elta, grand_total_edress,
                 grand_total_egross, grand_total_ot,grand_total_pot,grand_total_ppot,grand_total_bus, grand_total_fes, grand_total_arrear, grand_total_att, grand_total_supervisor,0,grand_total_shift,grand_total_spec, grand_total_etotal, 
                 grand_total_pf, grand_total_vpf, grand_total_totalpf, grand_total_elwf,grand_total_it, grand_total_pt, grand_total_advance,grand_total_mis, grand_total_esi, grand_total_totaldec, grand_total_dlop, grand_total_takehome])
    align_center = Alignment(horizontal='center',vertical='center')

    for cell in ws["1:1"]:
        cell.font = Font(bold=True,size=14)
        cell.alignment = align_center

    for cell in ws["2:2"]:
        if cell.column > 10 and cell.column != 50:
            # cell.font = Font(bold=True)
        # 	cell.alignment = align_center 
        # else:
            cell.font = Font(bold=True,size=10)
            cell.alignment = align_center     
    for cell in ws["2:2"]:
        cell.alignment = align_center 
    ws['A1'].fill = PatternFill(fgColor="FFFFFF", fill_type = "solid")

    for header in ws.iter_rows(min_row=2, max_row=3,min_col=1, max_col=52):
        for cell in header:
            cell.fill = PatternFill(fgColor='FFFF00', fill_type="solid")
    for header in ws.iter_rows(min_row=2, max_row=3,min_col=53, max_col=53):
        for cell in header:
            cell.fill = PatternFill(fgColor='FFC0CB', fill_type="solid")
    for header in ws.iter_rows(min_row=3, max_row=3,min_col=31, max_col=32):
        for cell in header:
            cell.fill = PatternFill(fgColor='3e8004', fill_type="solid")
    for header in ws.iter_rows(min_row=2, max_row=2,min_col=11, max_col=18):
        for cell in header:
            cell.fill = PatternFill(fgColor='CD853F', fill_type="solid")
            
    for header in ws.iter_rows(min_row=2, max_row=ws.max_row,min_col=19, max_col=21):
        for cell in header:
            cell.fill = PatternFill(fgColor='ADD8E6', fill_type="solid")
    
    border = Border(left=Side(border_style='thin', color='000000'),
                 right=Side(border_style='thin', color='000000'),
                 top=Side(border_style='thin', color='000000'),
                 bottom=Side(border_style='thin', color='000000'))

    for rows in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=53):
        for cell in rows:
            cell.border = border
    
    grey_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=11, max_col=18):
        for cell in row:
            cell.fill = grey_fill
    pink_fill = PatternFill(start_color='FFC0CB', end_color='FFC0CB', fill_type='solid')
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=53, max_col=53):
        for cell in row:
            cell.fill = pink_fill
    yellow_fill = PatternFill(start_color='fff600', end_color='fff600', fill_type='solid')
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=29, max_col=29):
        for cell in row:
            cell.fill = yellow_fill
    y_fill = PatternFill(start_color='fceda2', end_color='fceda2', fill_type='solid')
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=41, max_col=41):
        for cell in row:
            cell.fill = y_fill
    yellow = PatternFill(start_color='fff600',end_color='fff600',fill_type='solid')
    green = PatternFill(start_color='adff2f',end_color='adff2f',fill_type='solid')
    for rows in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=3, max_col=3):
        for cell in rows:
            if cell.value == 'Subtotal':
                for cell in ws[cell.row]:
                    cell.fill = green
                    ws.merge_cells(start_row=cell.row, start_column=2, end_row=cell.row, end_column=4) 
            if cell.value == 'Grand Total' or cell.value=='Total for Workers':
                for cell in ws[cell.row]:
                    cell.fill = yellow
                    ws.merge_cells(start_row=cell.row, start_column=2, end_row=cell.row, end_column=4) 

    max_row = ws.max_row
    previous_department = None
    merge_start_row = None

    for row in range(1, max_row + 1):
        cell_value = ws.cell(row=row, column=2).value

        # Apply center alignment to the current cell
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='center', vertical='center')
        
        if cell_value is None or cell_value.strip() == '':
            previous_department = None
            merge_start_row = None
        elif cell_value == previous_department:
            if merge_start_row is not None:
                end_row = row
                ws.merge_cells(start_row=merge_start_row, start_column=2, end_row=end_row, end_column=2)
        else:
            if merge_start_row is not None and previous_department is not None:
                end_row = row - 1
                ws.merge_cells(start_row=merge_start_row, start_column=2, end_row=end_row, end_column=2)
            
            previous_department = cell_value
            merge_start_row = row 

    # Handle the last group if needed
    if merge_start_row is not None and previous_department is not None:
        end_row = max_row
        ws.merge_cells(start_row=merge_start_row, start_column=2, end_row=end_row, end_column=2)
    for cell in ws["3:3"]:
        cell.alignment = align_center
    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    return xlsx_file

def build_xlsx_response(filename):
    xlsx_file = make_xlsx(filename)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'
    
