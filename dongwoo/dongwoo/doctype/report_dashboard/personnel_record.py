from __future__ import unicode_literals
import frappe
from frappe.utils import cstr, add_days, date_diff, getdate, format_date
from frappe import _, bold
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from frappe.utils.data import format_date
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue

from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook
import re
from frappe import _
import frappe
from frappe.model.document import Document
from datetime import date, timedelta, datetime,time
from frappe.utils import (getdate, cint, add_months, date_diff, add_days,
	nowdate, get_datetime_str, cstr, get_datetime, now_datetime, format_datetime,today, format_date)
import math
from frappe.utils import add_months, cint, flt, getdate, time_diff_in_hours,time_diff_in_seconds
import locale


import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types
import openpyxl.styles as styles

@frappe.whitelist()
def download():
	filename = 'Personnel Record'
	test = build_xlsx_response(filename)
	
def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
	args = frappe.local.form_dict
	column_widths = column_widths or []
	if wb is None:
		wb = openpyxl.Workbook()
		 
	ws = wb.create_sheet(sheet_name, 0)
	ws.column_dimensions['A'].width = 5
	ws.column_dimensions['B'].width = 10
	ws.column_dimensions['C'].width = 20 
	
	ws.append(title(args))
	ws.append(title1(args))
	ws.append(get_col(args))
	ws.append(get_col1(args))
	ws.append(get_col2(args))
	ws.append([''])
	data= get_data(args)
	for row in data:
		ws.append(row)
	
	ws.merge_cells(start_row=3, start_column=1, end_row=5, end_column=1 )
	ws.merge_cells(start_row=3, start_column=2, end_row=5, end_column=2 )
	ws.merge_cells(start_row=3, start_column=3, end_row=5, end_column=3 )
	ws.merge_cells(start_row=3, start_column=4, end_row=5, end_column=4 )
	ws.merge_cells(start_row=3, start_column=5, end_row=5, end_column=5 )
	ws.merge_cells(start_row=3, start_column=6, end_row=5, end_column=6 )
	ws.merge_cells(start_row=3, start_column=7, end_row=5, end_column=7 )
	ws.merge_cells(start_row=3, start_column=8, end_row=5, end_column=8 )
	ws.merge_cells(start_row=3, start_column=9, end_row=5, end_column=9 )
	ws.merge_cells(start_row=3, start_column=10, end_row=5, end_column=10 )
	ws.merge_cells(start_row=3, start_column=11, end_row=5, end_column=11 )
	ws.merge_cells(start_row=3, start_column=12, end_row=5, end_column=12 )
	year_data = get_yr()
	end_col = 13 + len(year_data) - 1

	ws.merge_cells(start_row=3, start_column=13, end_row=4, end_column=end_col )
	ws.merge_cells(start_row=3, start_column=end_col+1, end_row=5, end_column=end_col+1 )
	ws.merge_cells(start_row=3, start_column=end_col+2, end_row=4, end_column=end_col+len(year_data)+1 )
	ws.merge_cells(start_row=3, start_column=end_col+len(year_data)+2, end_row=4, end_column=end_col+len(year_data)+len(year_data)+1 )
	width=end_col+len(year_data)+len(year_data)+1
	ws.merge_cells(start_row=3, start_column=width+1, end_row=4, end_column=width+1 )
	ws.merge_cells(start_row=3, start_column=width+2, end_row=4, end_column=width+3 )
	ws.merge_cells(start_row=3, start_column=width+4, end_row=4, end_column=width+4)
	ws.merge_cells(start_row=3, start_column=width+5, end_row=5, end_column=width+ 5)
	ws.merge_cells(start_row=3, start_column=width+6, end_row=5, end_column=width+ 6)
	ws.merge_cells(start_row=3, start_column=width+7, end_row=5, end_column=width+7 )
	ws.merge_cells(start_row=3, start_column=width+8, end_row=4, end_column=width+9 )
	ws.merge_cells(start_row=3, start_column=width+10, end_row=4, end_column=width+11 )
	ws.merge_cells(start_row=3, start_column=width+12 , end_row=3, end_column=width+18 )
	ws.merge_cells(start_row=4, start_column=width+12, end_row=4, end_column=width+18)
	ws.merge_cells(start_row=3, start_column=width+19, end_row=5, end_column=width+19 )
	ws.merge_cells(start_row=3, start_column=width+20, end_row=5, end_column=width+20 )
	ws.merge_cells(start_row=3, start_column=width+21, end_row=5, end_column=width+21 )
	ws.merge_cells(start_row=3, start_column=width+22, end_row=5, end_column=width+22 )
	ws.merge_cells(start_row=3, start_column=width+23, end_row=5, end_column=width+23 )
	ws.merge_cells(start_row=3, start_column=width+24, end_row=5, end_column=width+24 )
	ws.merge_cells(start_row=3, start_column=width+25, end_row=5, end_column=width+25 )
	ws.merge_cells(start_row=3, start_column=width+26, end_row=5, end_column=width+26 )
	ws.merge_cells(start_row=3, start_column=width+26, end_row=5, end_column=width+26 )
	ws.merge_cells(start_row=3, start_column=width+27, end_row=5, end_column=width+27 )
	ws.merge_cells(start_row=3, start_column=width+28, end_row=5, end_column=width+28)
	ws.merge_cells(start_row=3, start_column=width+29, end_row=5, end_column=width+29 )
	ws.merge_cells(start_row=3, start_column=width+30, end_row=5, end_column=width+30 )
	ws.merge_cells(start_row=3, start_column=width+31, end_row=5, end_column=width+31 )
	to=width+32
	sal=frappe.db.get_all("Fiscal Year",["name"],order_by="name ASC")
	for s in sal:
		
		tot = to + 3
		ws.merge_cells(start_row=3, start_column= to, end_row=4, end_column= tot)
		to = tot + 1
	
	ws.merge_cells(start_row=3, start_column=width+width+38, end_row=4, end_column=width+width+41 )
	ws.merge_cells(start_row=3, start_column=width+width+42, end_row=4, end_column=width+width+44 )
	ws.merge_cells(start_row=3, start_column=width+width+45, end_row=5, end_column=width+width+45 )
	ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=width+width+45 )
	ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=width+width+45 )


	att=get_leave()
	leave=len(att)
	# ws.merge_cells(start_row=3, start_column=106, end_row=4, end_column=108 )
	# ws.merge_cells(start_row=2, start_column=109, end_row=4, end_column=109 )

	border_thin = Border(
	left=Side(style='thin'),
	right=Side(style='thin'),
	top=Side(style='thin'),
	bottom=Side(style='thin'))
	align_center = Alignment(horizontal='center',vertical='center')
	for header in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=1):
			for cell in header:
				cell.font = Font(bold=True,size=15)
	for header in ws.iter_rows(min_row=2, max_row=2, min_col=1, max_col=1):
			for cell in header:
				cell.font = Font(bold=True,size=10)
	for header in ws.iter_rows(min_row=3, max_row=6, min_col=1, max_col=+width+39+leave):
			for cell in header:
				cell.fill = PatternFill(fgColor='c7e9f0', fill_type = "solid")
				cell.font = Font(bold=True,size=10)
				cell.alignment = align_center
	header_range = ws['A1':ws.cell(row=len(get_data(args))+6, column=+width+39+leave).coordinate]
	for row in header_range:
		for cell in row:
			cell.border = border_thin
	ws.freeze_panes = 'D1'
	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	return xlsx_file

def build_xlsx_response(filename):
	xlsx_file = make_xlsx(filename)
	frappe.response['filename'] = filename + '.xlsx'
	frappe.response['filecontent'] = xlsx_file.getvalue()
	frappe.response['type'] = 'binary'

def title(args):
	et = args['employee_type']
	data=["PERSONNEL RECORD ("+ et +")"]
	return data

def title1(args):
	fdate = datetime.strptime(args['start_date'],'%Y-%m-%d')
	f_date = fdate.strftime('%Y.%m.%d')
	data=["Prepared on "+ f_date]
	return data

def get_yr():
	sal=frappe.db.get_all("Fiscal Year",["name"],order_by="name ASC")
	return sal
def get_leave():
	row=[]
	sal=frappe.db.get_all("Fiscal Year",["name"],order_by="name ASC")
	for i in sal:
		row.append("Attendance in "+i.name[0:4])
		row.append("")
		row.append("")
		row.append("")
	return row
def get_col(args):
	row=[]
	sal=frappe.db.get_all("Fiscal Year",["name"],order_by="name ASC")
	row += ["S.No.","Photo","Name","Department","Designation","Incharge","Employee ID","DOB","DOJ","Age","Years of Exp.","Working in DWSI","Salary (Gross in Rupees)"]
	for i in range(len(sal) - 1):
		row.append(sal[i].name)
	row+=["Joining Designation","Department Transference"]
	for i in range(len(sal) - 1):
		row.append(sal[i].name)
	row+=["Designation Transference"]
	for i in range(len(sal) - 1):
		row.append(sal[i].name)
	row+=["Previous Employeer","Qualification / Professional License","","Professional License","Present Address","Permanent Address","Location","Contact No.","","Email ID","","Family Details","","","","","","","Gender","Marital","Religion","Passport","Height (cm)","Weight (Kg)","Blood Group","Eyesight","Color-blindness","Hearing","Health Condition- Medical test- Medical history","Political Affairs","Individual Character"]
	for i in sal:
		row.append("Leave in "+i.name[0:4])
		row.append("")
		row.append("")
		row.append("")
	row+=["Disciplinary Record (If any)","","","","Award / Good Work Performance (If any)","","","Overseas Training (If any)"]
	return row
def get_col1(args):
	row=[]
	sal=frappe.db.get_all("Fiscal Year",["name"],order_by="name ASC")
	row +=["","","","","","","","","","","","",]
	for i in sal:
		row.append("") 
	row.append(" ")
	for i in sal:
		row.append("")
	for i in sal:
		row.append("")
	row += [""," "," "," ","","","","","","","","Name/DOB/Occupation/Coresidence","","","","","","","","","","","","","","","","","","",""]
	for i in sal:
		row.append("")
		row.append("")
		row.append("")
		row.append("")

	row+=["","","","","","","",""]
	return row
	
def get_col2(args):
	row=[]
	sal=frappe.db.get_all("Fiscal Year",["name"],order_by="name ASC")
	row +=["","","","","","","","","","","","",]
	for i in sal:
		row.append((i.name)) 
	row.append(" ")
	for i in sal:
		row.append(i.name)
	for i in sal:
		row.append(i.name)
	row += ["Latest / The before of the Latest","Regular Qualification (High rank 2 degree)","Additional Qualification (High rank 2 degree)","License Name / Issuing Organization / Aquisition yr.","","","","Official","Personal","Official","Personal","Father","Mother","Spouse","Child (1)","Child (2)","Child (3)","Child (4)","","","","","","","","","","","","",""]
	for i in sal:
		row.append("CL")
		row.append("SL")
		row.append("EL")
		row.append("Total")
	row+=["1st","2nd","3rd","4th","1st","2nd","3rd",""]
	return row

def get_data(args):
	data = []
	row = []
	ind = 1

	if args['employee_type']:
		emp = frappe.get_all("Employee",{'employee_type':args['employee_type'],'status':'Active'},['*'],order_by='name ASC')	
	else:
		emp = frappe.get_all("Employee",{'status':'Active'},['*'],order_by='name ASC')
	for e in emp:
		row = [ind,e.img,e.employee_name,e.department,e.designation,e.incharge,e.name,e.date_of_birth,e.date_of_joining,e.age,e.service_years_in_previous_companies,e.service_years_in_dwsi]
		fiscal_year = frappe.db.get_all("Fiscal Year",["name"],order_by="name ASC")
		for i in fiscal_year:
			sal = frappe.db.get_value("Salary Gross", {"parent": e.name,'periodin_yrs':i.name},['gross_salary'])
			if sal:
				# for g in sal:
				row.extend([sal])
			else:
				row.extend(["-"])
		row.append(e.joining_designation or '')
		for i in fiscal_year:
			dep = frappe.db.get_value("Department Change", {"parent": e.name,'period':i.name},['from_department'])
			if dep:
				row.extend([dep +" "+ i.name])
			else:
				row.extend([" "])
		for i in fiscal_year:
			# des=frappe.get_all("Designation Change",{"parent":e.name,"period":i.name},['*'])
			des = frappe.db.get_value("Designation Change", {"parent": e.name,'period':i.name},['from_designation'])
			if des:
				row.extend([des +" "+ i.name])
			else:
				row.extend([" "])


		pre_exp = frappe.get_all("Previous Work Experience", {"parent": e.name},['*'],order_by='from_date DESC', limit=1) 
		if pre_exp:
			for exp in pre_exp:
				row.append(exp.location)
		else:
			row.append(" ")
		qualification = frappe.get_all("Education", {"parent": e.name}, ['*'])
		regular = []
		additional = []

		if qualification:
			for q in qualification:
				if q.level == "Graduate" or q.level == "Under Graduate":
					regular.append(f"{q.qualification  or ' '} {q.year or ' '} {q.school_univ or ' '}")
					additional.append('')

				elif q.level == "Post Graduate":
					regular.append('')
					additional.append(f"{q.qualification  or ' '} {q.year or ' '} {q.school_univ or ' '}")

				else:
					regular.append('')
					additional.append('')

			row.extend(regular)
			row.extend(additional)
		
		else:
			row.extend([" ", " "])

		row.append(" ")	
		row.append(e.current_address or '')
		row.append(e.permanent_address or '')
		row.append(e.current_location or '')
		row.append(e.company_phone_number or '')
		row.append(e.personal_primary_contact_number or '')
		row.append(e.company_email or '')
		row.append(e.personal_email or '')
		row.append(e.father_detail_ or '')
		row.append(e.mother_details or '')
		row.append(e.spouse_details_ or '')
		row.append(e.child or '')
		row.append(e.child2 or '')
		row.append(e.child3 or '')
		row.append(e.child4 or '')
		row.append(e.gender or '')
		row.append(e.marital_status or '')
		row.append(e.religion or '')
		if e.passport_number___:
			passport="Yes"
		else:
			passport="No"
		row.append(passport)
		if e.height !=0:
			row.append(e.height)
		else:
			row.append("")
		if e.weight !=0:
			row.append(e.weight)
		else:
			row.append("")
		row.append(e.blood_group or '')
		row.append("")
		row.append("")
		row.append("")
		if e.self_health_issues_1!="None":
			row.append(e.self_health_issues_1 or '')
		else:
			row.append("")
		row.append(e.political_party_interested_in_political_party or '')
		row.append("")
		for i in fiscal_year:
			leave=frappe.get_all("Leave Details",{"parent":e.name,"Year":i.name},["*"])
			if leave:
				for l in leave:
					row.append(l.cl)
					row.append(l.sl)
					row.append(l.el)
					row.append(l.total)
			else:
				row+=["","","",""]
		warned= frappe.get_all("Disciplinary Action",{"parent":e.name},['*'],limit=4)
		if warned:
			if len(warned)!=4:
				warn=4-len(warned)
				wn=len(warned)+int(warn)
				for dis in range(wn):
					if dis < len(warned):
						row.append("Warning Ordered on " + str(warned[dis]['order_date']) + " for " + warned[dis]['reason'] or '')
					else:
						row.append("")
			else:
				for dis in warned:
					row+=["Warning Ordered on "+ str(dis.order_date)+" for "+dis.reason]
			# row.extend(discipline)
		else:
			row.append(" ")
		awards=frappe.get_all("Rewards And Recognition",{"parent":e.name},['*'],limit=3)
		if awards:
			if len(awards)!=3:
				award=3-len(awards)
				awd=len(awards)+int(award)
				for aw in range(awd):
					if aw < len(awards):
						row.append(awards[aw]['name_of_reward'] +" on "+str(awards[aw]["date_and_year"])+" for "+awards[aw]['reason'])
					else:
						row.append("")
			else:
				for aw in awards:
					row+=[aw.name_of_reward+" on "+str(aw.date_and_year)+" for "+aw.reason]
		else:
			row.append(" ")
		

		row.append(e.overseas_training_details or '')
		ind+=1
		
		data.append(row)
		# frappe.log_error(title='test', message=sal)

	
	return data