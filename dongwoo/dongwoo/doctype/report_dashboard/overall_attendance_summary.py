from __future__ import unicode_literals
import frappe
from frappe.utils import cstr, add_days, date_diff, getdate, format_date
from frappe import _, bold
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from frappe.utils.data import format_date
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue
from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook
import re
from frappe import _
import frappe
from frappe.model.document import Document
from frappe.utils import (getdate, cint, add_months, date_diff, add_days,
	nowdate, get_datetime_str, cstr, get_datetime, now_datetime, format_datetime,today, format_date)
import math
from frappe.utils import add_months, cint, flt, getdate, time_diff_in_hours,time_diff_in_seconds
import locale
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types
import openpyxl.styles as styles

@frappe.whitelist()
def download(start_date,end_date):
	filename = 'Overall Attendance Summary'
	args = {'start_date':start_date,'end_date':end_date}
	frappe.msgprint("Report is generating in the background,kindly check after few mins in the same page.")
	enqueue(overall_attendance_summary, queue='default', timeout=6000, event='overall_attendance_summary',filename=filename,args=args)
	# test = build_xlsx_response(filename,args)
	
def make_xlsx(data, args,sheet_name=None, wb=None, column_widths=None):
	column_widths = column_widths or []
	if wb is None:
		wb = openpyxl.Workbook()
		 
	ws = wb.create_sheet(sheet_name, 0)
	ws.column_dimensions['A'].width = 25
	ws.append(['DongWoo Surfacetech (India) Pvt Ltd.'])
	ws.append([''])
	date1 = datetime.strptime(args['start_date'],'%Y-%m-%d')
	date11 = date1.strftime('%B %d,%Y')
	date2 = datetime.strptime(args['end_date'],'%Y-%m-%d')
	date22 = date2.strftime('%B %d,%Y')
	ws.append(['Overall Attendance Summary - From ' +  date11 + " To " + date22])
	ws.append([''])
	ws.append(title1(args))
	ws.append(title2(args))
	data=get_data(args)
	for d in data:
		ws.append(d)
	ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=len(title1(args)) )
	ws.merge_cells(start_row=3, start_column=1, end_row=4, end_column=len(title1(args)) )
	ws.merge_cells(start_row=5, start_column=1, end_row=6, end_column=1 )
	dates = get_dates(args)
	to = 2
	for date in dates:
		dept=frappe.db.count("Department",{'name':('!=',"All Departments")})
		tot = to + dept + 1
		ws.merge_cells(start_row=5, start_column= to, end_row=5, end_column= tot)
		to = tot + 1
		
	border_thin = Border(
	left=Side(style='thin'),
	right=Side(style='thin'),
	top=Side(style='thin'),
	bottom=Side(style='thin'))

	for rows in ws.iter_rows(min_row=1, max_row=len(get_data(args))+6, min_col=1, max_col=len(title1(args))):
		for cell in rows:
			cell.border = border_thin  

	align_center = Alignment(horizontal='center',vertical='center')
	for header in ws.iter_rows(min_row=7, max_row=len(get_data(args))+6, min_col=1, max_col=1):
			for cell in header:
				cell.font = Font(bold=True,size=10)
				cell.alignment = align_center
	for header in ws.iter_rows(min_row=1, max_row=6, min_col=1, max_col=len(title1(args))):
			for cell in header:
				cell.font = Font(bold=True,size=10)
				cell.alignment = align_center
	for header in ws.iter_rows(min_row=len(get_data(args))+6, max_row=len(get_data(args))+6, min_col=1, max_col=len(title1(args))):
			for cell in header:
				cell.font = Font(bold=True,size=10)
				cell.alignment = align_center
	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	return xlsx_file

def overall_attendance_summary(filename,args):
	frappe.log_error(title='test',message=filename)
	xlsx_file = make_xlsx(filename,args)
	ret = frappe.get_doc({
			"doctype": "File",
			"attached_to_name": ' ',
			"attached_to_doctype": 'Report Dashboard',
			"attached_to_field": 'attach',
			"file_name": filename + '.xlsx',
			"is_private": 0,
			"content": xlsx_file.getvalue(),
			"decode": False
		})
	ret.save(ignore_permissions=True)
	frappe.db.commit()
	attached_file = frappe.get_doc("File", ret.name)
	frappe.db.set_value('Report Dashboard',None,'attach',attached_file.file_url)

@frappe.whitelist()
def title1(args):
	data = []
	data += ['Employee Type/Department']
	dates = get_dates(args)
	for date in dates:
		date = datetime.strptime(date,'%Y-%m-%d')
		date = date.strftime('%A,%B %d,%Y')
		data.extend([date])
		dept=frappe.db.get_all("Department",['name'],order_by="order_value ASC")
		for d in dept:
			if d['name'] != "All Departments":
				data+=[""]
		data.append("")
	return data

@frappe.whitelist()
def title2(args):
	data = []
	data += ['']
	dates = get_dates(args)
	for date in dates:
		dept=frappe.db.get_all("Department",['name'],order_by="order_value ASC")
		for d in dept:
			if d.name != "All Departments":
				data+=[d['name']]
		data.append("W'OF")
		data.append("LEAVE")
	return data

@frappe.whitelist()
def dep(args):
	data=[]
	dates = get_dates(args)
	for date in dates:
		dept=frappe.db.get_all("Department",['name'],order_by="order_value ASC")
		for d in dept:
			data.append(d['name'])
		data.append("W'OF")
		data.append("LEAVE")
	return data

@frappe.whitelist()
def get_data(args):
	data=[]
	row1=[]
	row1.append('Staff')
	dates = get_dates(args)
	for date in dates:
		dept=frappe.db.get_all("Department",['name'],order_by="order_value ASC")
		for d in dept:
			if d['name'] != "All Departments":
				att = frappe.db.sql("""
						SELECT count(*) as count 
						FROM `tabAttendance` 
						WHERE attendance_date = %s 
							AND docstatus != 2 
							AND employee_type = 'Staff'
							AND department = %s AND status = 'Present'
					""", (date,d['name']), as_dict=True)[0]
				row1.append(att['count'])
		employee_type='Staff'
		row1.append(holiday(date,employee_type))
		att = frappe.db.sql("""
					SELECT count(*) as count 
					FROM `tabAttendance` 
					WHERE attendance_date = %s 
						AND docstatus != 2 
						AND employee_type = 'Staff'
						AND status = 'On Leave'
				""", (date), as_dict=True)[0]
		row1.append(att['count'])
		
	row2=[]
	row2.append('Worker')
	for date in dates:
		dept=frappe.db.get_all("Department",['name'],order_by="order_value ASC")
		for d in dept:
			if d['name'] != "All Departments":
				att = frappe.db.sql("""
						SELECT count(*) as count 
						FROM `tabAttendance` 
						WHERE attendance_date = %s 
							AND docstatus != 2 
							AND employee_type = 'Worker'
							AND department = %s AND status = 'Present'
					""", (date,d['name']), as_dict=True)[0]
				row2.append(att['count'])
		employee_type='Worker'
		row2.append(holiday(date,employee_type))
		att = frappe.db.sql("""
						SELECT count(*) as count 
						FROM `tabAttendance` 
						WHERE attendance_date = %s 
							AND docstatus != 2 
							AND employee_type = 'Worker'
							AND status = 'On Leave'
					""", (date), as_dict=True)[0]
		row2.append(att['count'])
		
	
	row3=[]
	row3.append('D . Trainee')
	for date in dates:
		dept=frappe.db.get_all("Department",['name'],order_by="order_value ASC")
		for d in dept:
			if d['name'] != "All Departments":
				att = frappe.db.sql("""
						SELECT count(*) as count 
						FROM `tabAttendance` 
						WHERE attendance_date = %s 
							AND docstatus != 2 
							AND employee_type = 'D . Trainee'
							AND department = %s AND status = 'Present'
					""", (date,d['name']), as_dict=True)[0]
				row3.append(att['count'])
		employee_type='D . Trainee'
		row3.append(holiday(date,employee_type))
		att = frappe.db.sql("""
						SELECT count(*) as count 
						FROM `tabAttendance` 
						WHERE attendance_date = %s 
							AND docstatus != 2 
							AND employee_type = 'D . Trainee'
							AND status = 'On Leave'
					""", (date), as_dict=True)[0]
		row3.append(att['count'])

	row6=[]
	row6.append('NAPS')
	dates = get_dates(args)
	for date in dates:
		dept=frappe.db.get_all("Department",['name'],order_by="order_value ASC")
		for d in dept:
			if d['name'] != "All Departments":
				att = frappe.db.sql("""
						SELECT count(*) as count 
						FROM `tabAttendance` 
						WHERE attendance_date = %s 
							AND docstatus != 2 
							AND employee_type = 'NAPS'
							AND department = %s AND status = 'Present'
					""", (date,d['name']), as_dict=True)[0]
				row6.append(att['count'])
		employee_type='NAPS'
		row6.append(holiday(date,employee_type))
		att = frappe.db.sql("""
					SELECT count(*) as count 
					FROM `tabAttendance` 
					WHERE attendance_date = %s 
						AND docstatus != 2 
						AND employee_type = 'NAPS'
						AND status = 'On Leave'
				""", (date), as_dict=True)[0]
		row6.append(att['count'])
		
	row4=[]
	contractor=frappe.db.get_all("Contractor",['name'])
	for c in contractor:
		contractor =[c['name']]
		for date in dates:
			dept=frappe.db.get_all("Department",['name'],order_by="order_value ASC")
			for d in dept:
				if d['name'] != "All Departments":
					att = frappe.db.sql("""
							SELECT count(*) as count 
							FROM `tabAttendance` 
							WHERE attendance_date = %s 
								AND docstatus != 2 
								AND contractor = %s
								AND department = %s AND status = 'Present'
						""", (date,c['name'],d['name']), as_dict=True)[0]
					contractor.append(att['count'])
			employee_type='Contractor'
			contractor.append(holiday(date,employee_type))
			att = frappe.db.sql("""
						SELECT count(*) as count 
						FROM `tabAttendance` 
						WHERE attendance_date = %s 
							AND docstatus != 2 
							AND contractor = %s
							AND status = 'On Leave'
					""", (date,c['name']), as_dict=True)[0]
			contractor.append(att['count'])
		row4.append(contractor)
		
	row5=[]
	row5+=['Total']
	for date in dates:
		dept=frappe.db.get_all("Department",['name'],order_by="order_value ASC")
		for d in dept:
			if d['name'] != "All Departments":
				att = frappe.db.sql("""
					SELECT count(*) as count 
					FROM `tabAttendance` 
					WHERE attendance_date = %s 
						AND docstatus != 2 
						AND department = %s AND status = 'Present'
				""", (date,d['name']), as_dict=True)[0]
				row5.append(att['count'])
		row5.append(holiday_total(date))
		att = frappe.db.sql("""
						SELECT count(*) as count 
						FROM `tabAttendance` 
						WHERE attendance_date = %s 
							AND docstatus != 2 
							AND status = 'On Leave'
					""", (date), as_dict=True)[0]
		row5.append(att['count'])

	data.append(row1)
	data.append(row2)
	data.append(row3)
	data.append(row6)
	data.extend(row4)
	data.append(row5)
	return data


def get_dates(args):
	no_of_days = date_diff(add_days(args['end_date'], 1), args['start_date'])
	dates = [add_days(args['start_date'], i) for i in range(0, no_of_days)]
	return dates

def holiday(date, employee_type):
	left_employees = []
	employees = frappe.db.sql("""
		SELECT name, holiday_list, employee_type 
		FROM `tabEmployee` 
		WHERE status = 'Active' AND employee_type = %s
	""", (employee_type,), as_dict=True)
	left_employees = frappe.db.sql("""
		SELECT name, employee_type 
		FROM `tabEmployee`
		WHERE status = 'Left' AND relieving_date >= %s AND employee_type = %s
	""", (date, employee_type), as_dict=True)
	
	employees.extend(left_employees)
	status = 0
	for e in employees:
		frappe.errprint("hi")
		holiday_list = e.get('holiday_list')
		holiday = frappe.db.sql("""
			SELECT `tabHoliday`.holiday_date, `tabHoliday`.weekly_off 
			FROM `tabHoliday List` 
			LEFT JOIN `tabHoliday` ON `tabHoliday`.parent = `tabHoliday List`.name 
			WHERE `tabHoliday List`.name = %s AND holiday_date = %s
		""", (holiday_list, date), as_dict=True)

		
		if holiday and holiday[0].weekly_off == 1:
			status += 1
	return status

@frappe.whitelist()
def holiday_total(date):
	left_employees = []
	employees = frappe.db.sql("""
		SELECT name, holiday_list
		FROM `tabEmployee` 
		WHERE status = 'Active'
	""", as_dict=True)
	left_employees = frappe.db.sql("""
		SELECT name
		FROM `tabEmployee`
		WHERE status = 'Left' AND relieving_date >= %s
	""", (date), as_dict=True)
	
	employees.extend(left_employees)
	status = 0
	for e in employees:
		frappe.errprint("hi")
		holiday_list = e.get('holiday_list')
		holiday = frappe.db.sql("""
			SELECT `tabHoliday`.holiday_date, `tabHoliday`.weekly_off 
			FROM `tabHoliday List` 
			LEFT JOIN `tabHoliday` ON `tabHoliday`.parent = `tabHoliday List`.name 
			WHERE `tabHoliday List`.name = %s AND holiday_date = %s
		""", (holiday_list, date), as_dict=True)
		
		if holiday and holiday[0].weekly_off == 1:
			status += 1
	return status