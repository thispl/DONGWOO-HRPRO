from __future__ import unicode_literals
import frappe
from frappe.utils import cstr, add_days, date_diff, getdate, format_date
from frappe import _, bold
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from frappe.utils.data import format_date
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue
from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook
import re
from frappe import _
import frappe
from frappe.model.document import Document
from datetime import date, timedelta, datetime,time
from frappe.utils import (getdate, cint, add_months, date_diff, add_days,
	nowdate, get_datetime_str, cstr, get_datetime, now_datetime, format_datetime,today, format_date)
import math
from frappe.utils import add_months, cint, flt, getdate, time_diff_in_hours,time_diff_in_seconds
import locale
import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types
import openpyxl.styles as styles

@frappe.whitelist()
def download():
	start_date = add_days(today(), -1)
	filename = 'Attendance Summary {}'.format(start_date)
	args = {'start_date':add_days(today(),-1),'end_date':add_days(today(),-1)}
	build_xlsx_response(filename=filename,args=args)
	
def make_xlsx(data,args, sheet_name=None, wb=None, column_widths=None):
	column_widths = column_widths or []
	if wb is None:
		wb = openpyxl.Workbook()
		 
	ws = wb.create_sheet(sheet_name, 0)

	header_date = get_title()
	ws.append(header_date)

	header_date = get_title_1(args)
	ws.append(header_date)

	header_date = get_title_2()
	ws.append(header_date)

	header_date = get_title_3()
	ws.append(header_date)

	data=get_data_1(args)
	for d in data:
		ws.append(d)

	header_date = get_title_6()
	ws.append(header_date)

	header_date = get_title_4()
	ws.append(header_date) 

	header_date = get_title_3()
	ws.append(header_date)

	data=get_data_2(args)
	for d in data:
		ws.append(d)

	header_date = get_title_6()
	ws.append(header_date)

	header_date = get_title_5()
	ws.append(header_date)

	header_date = get_title_3()
	ws.append(header_date)

	data=get_data_3(args)
	for d in data:
		ws.append(d)

	ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column= len(get_title_3()) )
	ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column= len(get_title_3()) )

	ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column= 2 )
	ws.merge_cells(start_row=3, start_column=3, end_row=3, end_column= 2 + ((len(get_title_3()) - 2)/3) )
	ws.merge_cells(start_row=3, start_column=3 + ((len(get_title_3()) - 2)/3), end_row=3, end_column= 2 + ((len(get_title_3()) - 2)/3) + ((len(get_title_3()) - 2)/3) )
	ws.merge_cells(start_row=3, start_column=3 + ((len(get_title_3()) - 2)/3) + ((len(get_title_3()) - 2)/3), end_row=3, end_column= len(get_title_3()) )
	ws.merge_cells(start_row=len(get_data_1(args)) + 4, start_column=1, end_row=len(get_data_1(args)) + 4, end_column= 2)

	ws.merge_cells(start_row=len(get_data_1(args)) + 5, start_column=1, end_row=len(get_data_1(args)) + 5, end_column= len(get_title_3()))
	
	ws.merge_cells(start_row=len(get_data_1(args)) + 6, start_column=1, end_row=len(get_data_1(args)) + 6, end_column= 2 )
	ws.merge_cells(start_row=len(get_data_1(args)) + 6, start_column=3, end_row=len(get_data_1(args)) + 6, end_column= 2 + ((len(get_title_3()) - 2)/3) )
	ws.merge_cells(start_row=len(get_data_1(args)) + 6, start_column=3 + ((len(get_title_3()) - 2)/3), end_row=len(get_data_1(args)) + 6, end_column= 2 + ((len(get_title_3()) - 2)/3) + ((len(get_title_3()) - 2)/3) )
	ws.merge_cells(start_row=len(get_data_1(args)) + 6, start_column=3 + ((len(get_title_3()) - 2)/3) + ((len(get_title_3()) - 2)/3), end_row=len(get_data_1(args)) + 6, end_column= len(get_title_3()) )
	ws.merge_cells(start_row=len(get_data_1(args)) + 7 + len(get_data_1(args)), start_column=1, end_row=len(get_data_1(args)) + 7 + len(get_data_1(args)), end_column= 2)
	
	ws.merge_cells(start_row=len(get_data_1(args)) + 8 + len(get_data_1(args)), start_column=1, end_row=len(get_data_1(args)) + 8 + len(get_data_1(args)), end_column= len(get_title_3()) )
	
	ws.merge_cells(start_row=len(get_data_1(args)) + 9 + len(get_data_1(args)), start_column=1, end_row=len(get_data_1(args)) + 9 + len(get_data_1(args)), end_column= 2 )
	ws.merge_cells(start_row=len(get_data_1(args)) + 9 + len(get_data_1(args)), start_column=3, end_row=len(get_data_1(args)) + 9 + len(get_data_1(args)), end_column= 2 + ((len(get_title_3()) - 2)/3) )
	ws.merge_cells(start_row=len(get_data_1(args)) + 9 + len(get_data_1(args)), start_column=3 + ((len(get_title_3()) - 2)/3), end_row=len(get_data_1(args)) + 9 + len(get_data_1(args)), end_column= 2 + ((len(get_title_3()) - 2)/3) + ((len(get_title_3()) - 2)/3) )
	ws.merge_cells(start_row=len(get_data_1(args)) + 9 + len(get_data_1(args)), start_column=3 + ((len(get_title_3()) - 2)/3) + ((len(get_title_3()) - 2)/3), end_row=len(get_data_1(args)) + 9 + len(get_data_1(args)), end_column= len(get_title_3()) )
	ws.merge_cells(start_row=len(get_data_1(args)) + 10 + len(get_data_1(args)) + len(get_data_1(args)), start_column=1, end_row=len(get_data_1(args)) + 10 + len(get_data_1(args)) + len(get_data_1(args)), end_column= 2)
	
	align_center = Alignment(horizontal='center',vertical='center')
	left = Alignment(horizontal="left")
	border = Border(
		left=Side(border_style='thin'),
		right=Side(border_style='thin'),
		top=Side(border_style='thin'),
		bottom=Side(border_style='thin'))
	for rows in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=len(get_title_3())):
		for cell in rows:
			cell.font = Font(bold=True)
			cell.alignment = align_center
	for rows in ws.iter_rows(min_row=3, max_row=len(get_data_1(args)) + 10 + len(get_data_1(args)) + len(get_data_1(args)), min_col=1, max_col=len(get_title_3())):
		for cell in rows:
			cell.font = Font(bold=True)
			cell.alignment = align_center
	for rows in ws.iter_rows(min_row=3, min_col=1, max_row=len(get_data_1(args)) + 10 + len(get_data_1(args)) + len(get_data_1(args)), max_col= len(get_title_3()) ): 
		for cell in rows:
			cell.font = Font(bold=True)
			cell.fill = PatternFill(fgColor='DAD8D9', fill_type="solid")
			cell.alignment = align_center
	for rows in ws.iter_rows(min_row=3, min_col=3 + ((len(get_title_3()) - 2) // 3), max_row=len(get_data_1(args)) + 10 + len(get_data_1(args)) + len(get_data_1(args)) , max_col=2 + ((len(get_title_3()) - 2) // 3) + ((len(get_title_3()) - 2) // 3)):
		for cell in rows:
			cell.font = Font(bold=True)
			cell.fill = PatternFill(fgColor='FBF9FA', fill_type="solid")
			cell.alignment = align_center
	
	for rows in ws.iter_rows(min_row=1, min_col=1, max_row=len(get_data_1(args)) + 10 + len(get_data_1(args)) + len(get_data_1(args)), max_col= len(get_title_3()) ):
		for cell in rows:
			cell.font = Font(bold=True)
			cell.border = border

	for rows in ws.iter_rows(min_row=5, min_col=1, max_row= len(get_data_1(args)) + 3, max_col=2):
		for cell in rows:
			cell.alignment = left
			cell.font = Font(bold=False)
			cell.fill = PatternFill(fgColor='FBF9FA', fill_type="solid")
	
	for rows in ws.iter_rows(min_row=len(get_data_1(args)) + 8, min_col=1, max_row= len(get_data_1(args)) + 6 + len(get_data_1(args)), max_col=2):
		for cell in rows:
			cell.alignment = left
			cell.font = Font(bold=False)
			cell.fill = PatternFill(fgColor='FBF9FA', fill_type="solid")

	for rows in ws.iter_rows(min_row=len(get_data_1(args)) + 11 + len(get_data_1(args)), min_col=1, max_row= len(get_data_1(args)) + 9 + len(get_data_1(args)) + len(get_data_1(args)), max_col=2):
		for cell in rows:
			cell.alignment = left
			cell.font = Font(bold=False)
			cell.fill = PatternFill(fgColor='FBF9FA', fill_type="solid")

	ws.column_dimensions['A'].width = 20
	ws.column_dimensions['B'].width = 30

	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	return xlsx_file

def build_xlsx_response(filename,args):
	xlsx_file = make_xlsx(filename, args)
	ret = frappe.get_doc({
			"doctype": "File",
			"attached_to_name": ' ',
			"attached_to_doctype": 'Daily Attendance Summary',
			"attached_to_field": 'attach',
			"file_name": filename + '.xlsx',
			"is_private": 0,
			"content": xlsx_file.getvalue(),
			"decode": False
		})
	ret.save(ignore_permissions=True)
	frappe.db.commit()
	attached_file = frappe.get_doc("File", ret.name)
	frappe.db.set_value('Daily Attendance Summary',None,'attach',attached_file.file_url)


@frappe.whitelist()
def get_title():
	status = []
	status = ['DongWoo Surfacetech (India) Pvt Ltd']
	return status

@frappe.whitelist()
def get_title_1(args):
	status = []
	status = ['Attendance Summary ' + args['start_date']]
	return status

@frappe.whitelist()
def get_title_2():
	row = []
	row += ["Regular"," "]
	shift_types = frappe.db.sql("""SELECT * FROM `tabShift Type` ORDER BY `name` ASC""", as_dict=True)
	for shift_type in shift_types:
		count = frappe.db.count("Contractor")
		row += ["Shift " + shift_type.name]
		for _ in range(count +	4):
			row += [" "]
	return row

@frappe.whitelist()
def get_title_3():
	result = []
	result += ["Parent","Department"]
	shifts = frappe.db.sql("""SELECT * FROM `tabShift Type` ORDER BY `name` ASC""", as_dict=True)
	for shift in shifts:
		ec = frappe.db.sql("""SELECT * FROM `tabEmployee Type` ORDER BY `order` ASC""", as_dict=True)
		for e in ec:
			if e.name != "Contract Employee" and e.name != "Director":
				result += [e.name]
			elif e.name == "Contract Employee":
				contractors = frappe.db.sql("""SELECT * FROM `tabContractor` ORDER BY `name` ASC""", as_dict=True)
				for contractor in contractors:
					result += [contractor.name]
		result += ["Total"]
	return result

@frappe.whitelist()
def get_title_4():
	row = []
	row += ["Overtime"," "]
	shift_types = frappe.db.sql("""SELECT * FROM `tabShift Type` ORDER BY `name` ASC""", as_dict=True)
	for shift_type in shift_types:
		count = frappe.db.count("Contractor")
		count1 = frappe.db.count("Employee Type")
		row += ["Shift " + shift_type.name]
		for _ in range(count-2 +count1):
			row += [" "]
	return row

@frappe.whitelist()
def get_title_5():
	row = []
	row += ["Total (Regular + OT)"," "]
	shift_types = frappe.db.sql("""SELECT * FROM `tabShift Type` ORDER BY `name` ASC""", as_dict=True)
	for shift_type in shift_types:
		count = frappe.db.count("Contractor")
		count1 = frappe.db.count("Employee Type")
		row += ["Shift " + shift_type.name]
		for _ in range(count-2 +count1):
			row += [" "]
	return row

@frappe.whitelist()
def get_title_6():
	row = []
	row += [ ]
	shift_types = frappe.db.sql("""SELECT * FROM `tabShift Type` ORDER BY `name` ASC""", as_dict=True)
	for shift_type in shift_types:
		count = frappe.db.count("Contractor")
		count1 = frappe.db.count("Employee Type")
		row += [ ]
		for _ in range(count-2 +count1):
			row += [ ]
	return row


@frappe.whitelist()
def get_data_1(args):
	status = []
	departments = frappe.db.sql("""
		SELECT * FROM `tabDepartment` 
		WHERE name != "All Departments" AND is_group = 1 
		ORDER BY `name` ASC
	""", as_dict=True)
	for department in departments:
		sub_departments = frappe.db.get_all("Department", filters={'parent_department': department.name}, fields=['*'])
		for sub_department in sub_departments:
			row = [department.name, sub_department.name]
			shifts = frappe.db.sql("""SELECT * FROM `tabShift Type` ORDER BY `name` ASC""", as_dict=True)
			for shift in shifts:
				tot = 0
				ec = frappe.db.sql("""SELECT * FROM `tabEmployee Type` ORDER BY `order` ASC""", as_dict=True)
				for employee_type in ec:
					if employee_type.name != "Contract Employee" and employee_type.name != "Director":
						c = frappe.db.count("Attendance", {
							'attendance_date': (args['start_date']),
							'docstatus': ('!=', '2'),
							'shift': shift.name,
							'employee_type': employee_type.name,
							'department': sub_department.name,
							'status' : "Present"
						}) or 0
						tot += c
						if c > 0 :
							row.append(c)
						else :
							row.append('-')
					elif employee_type.name == "Contract Employee":
						contractors = frappe.db.sql("""SELECT * FROM `tabContractor` ORDER BY `name` ASC""", as_dict=True)
						for contractor in contractors:
							c = frappe.db.count("Attendance", {
								'attendance_date': (args['start_date']),
								'docstatus': ('!=', '2'),
								'shift': shift.name,
								'employee_type': employee_type.name,
								'department': sub_department.name,
								'status' : "Present",
								'contractor': contractor.name
							}) or 0
							tot += c
							if c > 0 :
								row.append(c)
							else :
								row.append('-')		
				row.append(tot)
			status.append(row)	
	row = ["Total",'']
	shifts = frappe.db.sql("""SELECT * FROM `tabShift Type` ORDER BY `name` ASC""", as_dict=True)
	for shift in shifts:
		tot = 0
		ec = frappe.db.sql("""SELECT * FROM `tabEmployee Type` ORDER BY `order` ASC""", as_dict=True)
		for employee_type in ec:
			if employee_type.name != "Contract Employee" and employee_type.name != "Director":
				c = frappe.db.count("Attendance", {
					'attendance_date': (args['start_date']),
					'docstatus': ('!=', '2'),
					'shift': shift.name,
					'employee_type': employee_type.name,
					'status' : "Present",

				}) or 0
				tot += c
				row.append(c)
			elif employee_type.name == "Contract Employee":

				contractors = frappe.db.sql("""SELECT * FROM `tabContractor` ORDER BY `name` ASC""", as_dict=True)
				for contractor in contractors:
					c = frappe.db.count("Attendance", {
						'attendance_date': (args['start_date']),
						'docstatus': ('!=', '2'),
						'shift': shift.name,
						'employee_type': employee_type.name,
						'status' : "Present",
						'contractor': contractor.name,
	
					}) or 0
					tot += c
					row.append(c)	
		row.append(tot)
	status.append(row)	
	return status


@frappe.whitelist()
def get_data_2(args):
	status = []
	departments = frappe.db.sql("""
		SELECT * FROM `tabDepartment` 
		WHERE name != "All Departments" AND is_group = 1 
		ORDER BY `name` ASC
	""", as_dict=True)
	for department in departments:
		sub_departments = frappe.db.get_all("Department", filters={'parent_department': department.name}, fields=['*'])
		for sub_department in sub_departments:
			row = [department.name, sub_department.name]
			shifts = frappe.db.sql("""SELECT * FROM `tabShift Type` ORDER BY `name` ASC""", as_dict=True)
			for shift in shifts:
				tot = 0
				ec = frappe.db.sql("""SELECT * FROM `tabEmployee Type` ORDER BY `order` ASC""", as_dict=True)
				for employee_type in ec:
					if employee_type.name != "Contract Employee" and employee_type.name != "Director":
						if shift.name == "A":
							b_shift_ot = frappe.db.count("Attendance", {
								'attendance_date': add_days((args['start_date']),-1),
								'docstatus': ('!=', '2'),
								'shift': "B",
								'employee_type': employee_type.name,
								'department': sub_department.name,
								'overtime_hours': ('>=', 15)
							}) or 0
							c_shift_ot = frappe.db.count("Attendance", {
								'attendance_date': add_days((args['start_date']),-1),
								'docstatus': ('!=', '2'),
								'shift': "C",
								'employee_type': employee_type.name,
								'department': sub_department.name,
								'overtime_hours': ('>=', 8)
							}) or 0
							ot = b_shift_ot + c_shift_ot
						if shift.name == "B":
							c_shift_ot = frappe.db.count("Attendance", {
								'attendance_date': add_days((args['start_date']),-1),
								'docstatus': ('!=', '2'),
								'shift': "C",
								'employee_type': employee_type.name,
								'department': sub_department.name,
								'overtime_hours': ('>=', 16)
							}) or 0
							a_shift_ot = frappe.db.count("Attendance", {
								'attendance_date': add_days((args['start_date']),0),
								'docstatus': ('!=', '2'),
								'shift': "A",
								'employee_type': employee_type.name,
								'department': sub_department.name,
								'overtime_hours': ('>=', 8)
							}) or 0
							ot = a_shift_ot + c_shift_ot
						if shift.name == "C":
							a_shift_ot = frappe.db.count("Attendance", {
								'attendance_date': add_days((args['start_date']),0),
								'docstatus': ('!=', '2'),
								'shift': "A",
								'employee_type': employee_type.name,
								'department': sub_department.name,
								'overtime_hours': ('>=', 15)
							})
							b_shift_ot = frappe.db.count("Attendance", {
								'attendance_date': add_days((args['start_date']),0),
								'docstatus': ('!=', '2'),
								'shift': "B",
								'employee_type': employee_type.name,
								'department': sub_department.name,
								'overtime_hours': ('>=', 7)
							})
							ot = a_shift_ot + b_shift_ot
						tot += ot
						if ot > 0 :
							row.append(ot)
						else:
							row.append('-')
					elif employee_type.name == "Contract Employee":
						contractors = frappe.db.sql("""SELECT * FROM `tabContractor` ORDER BY `name` ASC""", as_dict=True)
						for contractor in contractors:
							if shift.name == "A":
								b_shift_ot = frappe.db.count("Attendance", {
									'attendance_date': add_days((args['start_date']),-1),
									'docstatus': ('!=', '2'),
									'shift': "B",
									'employee_type': employee_type.name,
									'department': sub_department.name,
									'contractor': contractor.name,
									'overtime_hours': ('>=', 15)
								}) or 0
								c_shift_ot = frappe.db.count("Attendance", {
									'attendance_date': add_days((args['start_date']),-1),
									'docstatus': ('!=', '2'),
									'shift': "C",
									'employee_type': employee_type.name,
									'department': sub_department.name,
									'contractor': contractor.name,
									'overtime_hours': ('>=', 8)
								}) or 0
								ot = b_shift_ot + c_shift_ot
							if shift.name == "B":
								c_shift_ot = frappe.db.count("Attendance", {
									'attendance_date': add_days((args['start_date']),-1),
									'docstatus': ('!=', '2'),
									'shift': "C",
									'employee_type': employee_type.name,
									'department': sub_department.name,
									'contractor': contractor.name,
									'overtime_hours': ('>=', 16)
								}) or 0
								a_shift_ot = frappe.db.count("Attendance", {
									'attendance_date': add_days((args['start_date']),0),
									'docstatus': ('!=', '2'),
									'shift': "A",
									'employee_type': employee_type.name,
									'department': sub_department.name,
									'contractor': contractor.name,
									'overtime_hours': ('>=', 8)
								}) or 0
								ot = a_shift_ot + c_shift_ot
							if shift.name == "C":
								a_shift_ot = frappe.db.count("Attendance", {
									'attendance_date': add_days((args['start_date']),0),
									'docstatus': ('!=', '2'),
									'shift': "A",
									'employee_type': employee_type.name,
									'department': sub_department.name,
									'contractor': contractor.name,
									'overtime_hours': ('>=', 15)
								}) or 0
								b_shift_ot = frappe.db.count("Attendance", {
									'attendance_date': add_days((args['start_date']),0),
									'docstatus': ('!=', '2'),
									'shift': "B",
									'employee_type': employee_type.name,
									'department': sub_department.name,
									'contractor': contractor.name,
									'overtime_hours': ('>=', 7)
								}) or 0
								ot = a_shift_ot + b_shift_ot
							tot += ot
							if ot > 0 :
								row.append(ot)
							else:
								row.append('-')		
				row.append(tot)
			status.append(row)
	row = ["Total",'']
	shifts = frappe.db.sql("""SELECT * FROM `tabShift Type` ORDER BY `name` ASC""", as_dict=True)
	for shift in shifts:
		tot = 0
		ec = frappe.db.sql("""SELECT * FROM `tabEmployee Type` ORDER BY `order` ASC""", as_dict=True)
		for employee_type in ec:
			if employee_type.name != "Contract Employee" and employee_type.name != "Director":
				if shift.name == "A":
					b_shift_ot = frappe.db.count("Attendance", {
						'attendance_date': add_days((args['start_date']),-1),
						'docstatus': ('!=', '2'),
						'shift': "B",
						'employee_type': employee_type.name,
						'overtime_hours': ('>=', 15)
					}) or 0
					c_shift_ot = frappe.db.count("Attendance", {
						'attendance_date': add_days((args['start_date']),-1),
						'docstatus': ('!=', '2'),
						'shift': "C",
						'employee_type': employee_type.name,
						'overtime_hours': ('>=', 8)
					}) or 0
					ot = b_shift_ot + c_shift_ot
				if shift.name == "B":
					c_shift_ot = frappe.db.count("Attendance", {
						'attendance_date': add_days((args['start_date']),-1),
						'docstatus': ('!=', '2'),
						'shift': "C",
						'employee_type': employee_type.name,
						'overtime_hours': ('>=', 16)
					}) or 0
					a_shift_ot = frappe.db.count("Attendance", {
						'attendance_date': add_days((args['start_date']),0),
						'docstatus': ('!=', '2'),
						'shift': "A",
						'employee_type': employee_type.name,
						'overtime_hours': ('>=', 8)
					}) or 0
					ot = a_shift_ot + c_shift_ot
				if shift.name == "C":
					a_shift_ot = frappe.db.count("Attendance", {
						'attendance_date': add_days((args['start_date']),0),
						'docstatus': ('!=', '2'),
						'shift': "A",
						'employee_type': employee_type.name,
						'overtime_hours': ('>=', 15)
					}) or 0
					b_shift_ot = frappe.db.count("Attendance", {
						'attendance_date': add_days((args['start_date']),0),
						'docstatus': ('!=', '2'),
						'shift': "B",
						'employee_type': employee_type.name,
						'overtime_hours': ('>=', 7)
					}) or 0
					ot = a_shift_ot + b_shift_ot
				tot += ot
				row.append(ot)
			elif employee_type.name == "Contract Employee":
				contractors = frappe.db.sql("""SELECT * FROM `tabContractor` ORDER BY `name` ASC""", as_dict=True)
				for contractor in contractors:
					if shift.name == "A":
						b_shift_ot = frappe.db.count("Attendance", {
							'attendance_date': add_days((args['start_date']),-1),
							'docstatus': ('!=', '2'),
							'shift': "B",
							'employee_type': employee_type.name,
							'contractor': contractor.name,
							'overtime_hours': ('>=', 15)
						}) or 0
						c_shift_ot = frappe.db.count("Attendance", {
							'attendance_date': add_days((args['start_date']),-1),
							'docstatus': ('!=', '2'),
							'shift': "C",
							'employee_type': employee_type.name,
							'contractor': contractor.name,
							'overtime_hours': ('>=', 8)
						}) or 0
						ot = b_shift_ot + c_shift_ot
					if shift.name == "B":
						c_shift_ot = frappe.db.count("Attendance", {
							'attendance_date': add_days((args['start_date']),-1),
							'docstatus': ('!=', '2'),
							'shift': "C",
							'employee_type': employee_type.name,
							'contractor': contractor.name,
							'overtime_hours': ('>=', 16)
						}) or 0
						a_shift_ot = frappe.db.count("Attendance", {
							'attendance_date': add_days((args['start_date']),0),
							'docstatus': ('!=', '2'),
							'shift': "A",
							'employee_type': employee_type.name,
							'contractor': contractor.name,
							'overtime_hours': ('>=', 8)
						}) or 0
						ot = a_shift_ot + c_shift_ot
					if shift.name == "C":
						a_shift_ot = frappe.db.count("Attendance", {
							'attendance_date': add_days((args['start_date']),0),
							'docstatus': ('!=', '2'),
							'shift': "A",
							'employee_type': employee_type.name,
							'contractor': contractor.name,
							'overtime_hours': ('>=', 15)
						}) or 0
						b_shift_ot = frappe.db.count("Attendance", {
							'attendance_date': add_days((args['start_date']),0),
							'docstatus': ('!=', '2'),
							'shift': "B",
							'employee_type': employee_type.name,
							'contractor': contractor.name,
							'overtime_hours': ('>=', 7)
						}) or 0
						ot = a_shift_ot + b_shift_ot
					tot += ot
					row.append(ot)
		row.append(tot)
	status.append(row)		
	return status

@frappe.whitelist()
def get_data_3(args):
	status = []
	departments = frappe.db.sql("""SELECT * FROM `tabDepartment` WHERE name != "All Departments" AND is_group = 1 ORDER BY `name` ASC""", as_dict=True)
	for department in departments:
		sub_departments = frappe.db.get_all("Department", filters={'parent_department': department.name}, fields=['*'])
		for sub_department in sub_departments:
			row = [department.name, sub_department.name]
			shifts = frappe.db.sql("""SELECT * FROM `tabShift Type` ORDER BY `name` ASC""", as_dict=True)
			for shift in shifts:
				tot = 0
				ec = frappe.db.sql("""SELECT * FROM `tabEmployee Type` ORDER BY `order` ASC""", as_dict=True)
				for employee_type in ec:
					if employee_type.name != "Contract Employee" and employee_type.name != "Director":
						c = frappe.db.count("Attendance", {
							'attendance_date': (args['start_date']),
							'docstatus': ('!=', '2'),
							'shift': shift.name,
							'employee_type': employee_type.name,
							'department': sub_department.name,
							'status' : "Present"
						}) or 0
						tot += c
						if shift.name == "A":
							b_shift_ot = frappe.db.count("Attendance", {
								'attendance_date': add_days((args['start_date']),-1),
								'docstatus': ('!=', '2'),
								'shift': "B",
								'employee_type': employee_type.name,
								'department': sub_department.name,
								'overtime_hours': ('>=', 15)
							}) or 0
							c_shift_ot = frappe.db.count("Attendance", {
								'attendance_date': add_days((args['start_date']),-1),
								'docstatus': ('!=', '2'),
								'shift': "C",
								'employee_type': employee_type.name,
								'department': sub_department.name,
								'overtime_hours': ('>=', 8)
							}) or 0
							ot = b_shift_ot + c_shift_ot
						if shift.name == "B":
							c_shift_ot = frappe.db.count("Attendance", {
								'attendance_date': add_days((args['start_date']),-1),
								'docstatus': ('!=', '2'),
								'shift': "C",
								'employee_type': employee_type.name,
								'department': sub_department.name,
								'overtime_hours': ('>=', 16)
							}) or 0
							a_shift_ot = frappe.db.count("Attendance", {
								'attendance_date': add_days((args['start_date']),0),
								'docstatus': ('!=', '2'),
								'shift': "A",
								'employee_type': employee_type.name,
								'department': sub_department.name,
								'overtime_hours': ('>=', 8)
							}) or 0
							ot = a_shift_ot + c_shift_ot
						if shift.name == "C":
							a_shift_ot = frappe.db.count("Attendance", {
								'attendance_date': add_days((args['start_date']),0),
								'docstatus': ('!=', '2'),
								'shift': "A",
								'employee_type': employee_type.name,
								'department': sub_department.name,
								'overtime_hours': ('>=', 15)
							})
							b_shift_ot = frappe.db.count("Attendance", {
								'attendance_date': add_days((args['start_date']),0),
								'docstatus': ('!=', '2'),
								'shift': "B",
								'employee_type': employee_type.name,
								'department': sub_department.name,
								'overtime_hours': ('>=', 7)
							})
							ot = a_shift_ot + b_shift_ot
						tot += ot
						if (c + ot) > 0 :
							row.append((c + ot))
						else:
							row.append('-')
					elif employee_type.name == "Contract Employee":
						contractors = frappe.db.sql("""SELECT * FROM `tabContractor` ORDER BY `name` ASC""", as_dict=True)
						for contractor in contractors:
							c = frappe.db.count("Attendance", {
								'attendance_date': (args['start_date']),
								'docstatus': ('!=', '2'),
								'shift': shift.name,
								'employee_type': employee_type.name,
								'department': sub_department.name,
								'status' : "Present",
								'contractor': contractor.name
							}) or 0
							tot += c
							if shift.name == "A":
								b_shift_ot = frappe.db.count("Attendance", {
									'attendance_date': add_days((args['start_date']),-1),
									'docstatus': ('!=', '2'),
									'shift': "B",
									'employee_type': employee_type.name,
									'department': sub_department.name,
									'contractor': contractor.name,
									'overtime_hours': ('>=', 15)
								}) or 0
								c_shift_ot = frappe.db.count("Attendance", {
									'attendance_date': add_days((args['start_date']),-1),
									'docstatus': ('!=', '2'),
									'shift': "C",
									'employee_type': employee_type.name,
									'department': sub_department.name,
									'contractor': contractor.name,
									'overtime_hours': ('>=', 8)
								}) or 0
								ot = b_shift_ot + c_shift_ot
							if shift.name == "B":
								c_shift_ot = frappe.db.count("Attendance", {
									'attendance_date': add_days((args['start_date']),-1),
									'docstatus': ('!=', '2'),
									'shift': "C",
									'employee_type': employee_type.name,
									'department': sub_department.name,
									'contractor': contractor.name,
									'overtime_hours': ('>=', 16)
								}) or 0
								a_shift_ot = frappe.db.count("Attendance", {
									'attendance_date': add_days((args['start_date']),0),
									'docstatus': ('!=', '2'),
									'shift': "A",
									'employee_type': employee_type.name,
									'department': sub_department.name,
									'contractor': contractor.name,
									'overtime_hours': ('>=', 8)
								}) or 0
								ot = a_shift_ot + c_shift_ot
							if shift.name == "C":
								a_shift_ot = frappe.db.count("Attendance", {
									'attendance_date': add_days((args['start_date']),0),
									'docstatus': ('!=', '2'),
									'shift': "A",
									'employee_type': employee_type.name,
									'department': sub_department.name,
									'contractor': contractor.name,
									'overtime_hours': ('>=', 15)
								}) or 0
								b_shift_ot = frappe.db.count("Attendance", {
									'attendance_date': add_days((args['start_date']),0),
									'docstatus': ('!=', '2'),
									'shift': "B",
									'employee_type': employee_type.name,
									'department': sub_department.name,
									'contractor': contractor.name,
									'overtime_hours': ('>=', 7)
								}) or 0
								ot = a_shift_ot + b_shift_ot
							tot += (ot + c)
							if (ot + c) > 0 :
								row.append((ot + c))
							else:
								row.append('-')		
				row.append(tot)
			status.append(row)
	row = ["Total",'']
	shifts = frappe.db.sql("""SELECT * FROM `tabShift Type` ORDER BY `name` ASC""", as_dict=True)
	for shift in shifts:
		tot = 0
		ec = frappe.db.sql("""SELECT * FROM `tabEmployee Type` ORDER BY `order` ASC""", as_dict=True)
		for employee_type in ec:
			if employee_type.name != "Contract Employee" and employee_type.name != "Director":
				c = frappe.db.count("Attendance", {
					'attendance_date': (args['start_date']),
					'docstatus': ('!=', '2'),
					'shift': shift.name,
					'employee_type': employee_type.name,
					'status' : "Present"
				}) or 0
				tot += c
				if shift.name == "A":
					b_shift_ot = frappe.db.count("Attendance", {
						'attendance_date': add_days((args['start_date']),-1),
						'docstatus': ('!=', '2'),
						'shift': "B",
						'employee_type': employee_type.name,
						'overtime_hours': ('>=', 15)
					}) or 0
					c_shift_ot = frappe.db.count("Attendance", {
						'attendance_date': add_days((args['start_date']),-1),
						'docstatus': ('!=', '2'),
						'shift': "C",
						'employee_type': employee_type.name,
						'overtime_hours': ('>=', 8)
					}) or 0
					ot = b_shift_ot + c_shift_ot
				if shift.name == "B":
					c_shift_ot = frappe.db.count("Attendance", {
						'attendance_date': add_days((args['start_date']),-1),
						'docstatus': ('!=', '2'),
						'shift': "C",
						'employee_type': employee_type.name,
						'overtime_hours': ('>=', 16)
					}) or 0
					a_shift_ot = frappe.db.count("Attendance", {
						'attendance_date': add_days((args['start_date']),0),
						'docstatus': ('!=', '2'),
						'shift': "A",
						'employee_type': employee_type.name,
						'overtime_hours': ('>=', 8)
					}) or 0
					ot = a_shift_ot + c_shift_ot
				if shift.name == "C":
					a_shift_ot = frappe.db.count("Attendance", {
						'attendance_date': add_days((args['start_date']),0),
						'docstatus': ('!=', '2'),
						'shift': "A",
						'employee_type': employee_type.name,
						'overtime_hours': ('>=', 15)
					})
					b_shift_ot = frappe.db.count("Attendance", {
						'attendance_date': add_days((args['start_date']),0),
						'docstatus': ('!=', '2'),
						'shift': "B",
						'employee_type': employee_type.name,
						'overtime_hours': ('>=', 7)
					})
					ot = a_shift_ot + b_shift_ot
				tot += ot
				if (c + ot) > 0 :
					row.append((c + ot))
				else:
					row.append('-')
			elif employee_type.name == "Contract Employee":
				contractors = frappe.db.sql("""SELECT * FROM `tabContractor` ORDER BY `name` ASC""", as_dict=True)
				for contractor in contractors:
					c = frappe.db.count("Attendance", {
						'attendance_date': (args['start_date']),
						'docstatus': ('!=', '2'),
						'shift': shift.name,
						'employee_type': employee_type.name,
						'status' : "Present",
						'contractor': contractor.name
					}) or 0
					tot += c
					if shift.name == "A":
						b_shift_ot = frappe.db.count("Attendance", {
							'attendance_date': add_days((args['start_date']),-1),
							'docstatus': ('!=', '2'),
							'shift': "B",
							'employee_type': employee_type.name,
							'contractor': contractor.name,
							'overtime_hours': ('>=', 15)
						}) or 0
						c_shift_ot = frappe.db.count("Attendance", {
							'attendance_date': add_days((args['start_date']),-1),
							'docstatus': ('!=', '2'),
							'shift': "C",
							'employee_type': employee_type.name,
							'contractor': contractor.name,
							'overtime_hours': ('>=', 8)
						}) or 0
						ot = b_shift_ot + c_shift_ot
					if shift.name == "B":
						c_shift_ot = frappe.db.count("Attendance", {
							'attendance_date': add_days((args['start_date']),-1),
							'docstatus': ('!=', '2'),
							'shift': "C",
							'employee_type': employee_type.name,
							'contractor': contractor.name,
							'overtime_hours': ('>=', 16)
						}) or 0
						a_shift_ot = frappe.db.count("Attendance", {
							'attendance_date': add_days((args['start_date']),0),
							'docstatus': ('!=', '2'),
							'shift': "A",
							'employee_type': employee_type.name,
							'contractor': contractor.name,
							'overtime_hours': ('>=', 8)
						}) or 0
						ot = a_shift_ot + c_shift_ot
					if shift.name == "C":
						a_shift_ot = frappe.db.count("Attendance", {
							'attendance_date': add_days((args['start_date']),0),
							'docstatus': ('!=', '2'),
							'shift': "A",
							'employee_type': employee_type.name,
							'contractor': contractor.name,
							'overtime_hours': ('>=', 15)
						}) or 0
						b_shift_ot = frappe.db.count("Attendance", {
							'attendance_date': add_days((args['start_date']),0),
							'docstatus': ('!=', '2'),
							'shift': "B",
							'employee_type': employee_type.name,
							'contractor': contractor.name,
							'overtime_hours': ('>=', 7)
						}) or 0
						ot = a_shift_ot + b_shift_ot
					tot += (ot + c)
					if (ot + c) > 0 :
						row.append((ot + c))
					else:
						row.append('-')		
		row.append(tot)
	status.append(row)
	return status 

@frappe.whitelist()
def create_hooks_att3():
	job = frappe.db.exists('Scheduled Job Type', 'download')
	if not job:
		att = frappe.new_doc("Scheduled Job Type")
		att.update({
			"method": 'dongwoo.attendance_summary.download',
			"frequency": 'Cron',
			"cron_format": "*/5 * * * *"
		})
		att.save(ignore_permissions=True)
		
@frappe.whitelist()
def send_mail1():
	# attach1=frappe.db.get_value("Daily Attendance Summary",['attach'])
	attach1=frappe.get_doc('Daily Attendance Summary','Daily Attendance Summary')
	attachments='http://192.168.11.20'+attach1.attach
    
	# attachments = [{
	# 		'fname': filename,
	# 		'fcontent': content
	# 	}]
	formatted_date = add_days(today(),-1)
	date_obj = datetime.strptime(formatted_date, '%Y-%m-%d')
	formatted_date = date_obj.strftime('%d-%m-%Y')
	subject = f"Attendance Summary {formatted_date}"
	message = f"Dear Team,<br><br>Kindly find the attached Attendance Summary {formatted_date}."
	frappe.sendmail(
		# recipients= ['abdulla.pi@groupteampro.com','dineshbabu.k@groupteampro.com','anil.p@groupteampro.com','gifty.p@groupteampro.com',"venkatrajr@dwsi.co.in","vinothkumar@dwsi.co.in","vishnu@dwsi.co.in","security@dwsi.co.in"],
		recipients= ['gifty.p@groupteampro.com'],
		subject=subject,
		attachments=attachments,
		message=message
	)